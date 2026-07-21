"""
JKLC Battery Disconnection Mail Creation — web app adaptation (Report 6 of 9)
================================================================================

Adapted from the provided `jklc_battery_disconnection_mail.py` script and its
companion reference doc. Rebuilds the daily "Battery_Disconnection_Master_
<date>.xlsx" workbook Khagash maintains by hand -- this is a STATEFUL report:
each run needs the PREVIOUS day's Master workbook as an input (uploaded fresh
each time, same as any other file here -- there's no server-side persistence
between runs, the "state" is simply whatever file the user uploads).

Confirmed business logic (validated against real data -- see the reference
doc for the original evidence trail, plus independent re-verification below):

  1. Join key: 'Shipment No.' (consolidated report) == 'Shipment Number' (MTR).
  2. 'Consolidated' tab = pure unfiltered daily append log, forever (Clinker
     rows and same-day duplicate pings included).
  3. 'Consolidated Shipment No.' tab = same daily rows, per day: join Product
     Name from MTR -> drop Clinker rows -> dedupe by Shipment No. (keep first
     row in file order) -> STO/NON STO, Onward Status, shareable link,
     Test -> leave Mail Status / vehicle test blank for new rows, preserve
     existing values for carried-forward rows.
  4. 'Master' tab = shipments whose 'vehicle test' starts with "offline"
     (case-insensitive), append-only, excluding shipments whose latest
     'Mail Status' is "Waived OFF".
  5. 'MTR' tab = wholesale replaced with the latest MTR export each run,
     column order UNCHANGED from the raw export (required -- see below).
  6. 'Table' tab = carried forward unchanged (scratch/lookup helper,
     unrelated to the pipeline).

CELL-FOR-CELL FORMULA/FORMATTING FIDELITY (added after comparing our output
against the real reference workbook cell-by-cell, not just its computed
values):
  - STO/NON STO, Onward Status, and shareable link (Consolidated Shipment
    No. tab) and every Master column except 'Shipment No.'/'Freight loss'
    are live Excel formulas in the real workbook (an IFS + several
    XLOOKUPs), NOT static values -- see FORMULA_SPECS below. They're
    written as real formula cells here too, computed fresh for every row
    at write time rather than approximated in pandas, which also sidesteps
    a subtle mismatch: the real "Onward Status" formula actually reads
    MTR's "Stamp Status" column (not MTR's own "Onward Status" column,
    despite the CSN column's label) -- something a pandas .map() against
    the wrong-but-same-named column would never replicate.
  - Because those formulas hardcode MTR column LETTERS (e.g. "MTR!CI:CI"),
    the MTR tab's column order must exactly match the raw export's own
    order -- _load_mtr_lookup returns the untouched raw frame for this
    reason; do not reindex/reorder it before writing.
  - Every data cell (not just headers) on Consolidated/Consolidated
    Shipment No./Master uses "Aptos Narrow" 10pt + thin border on all 4
    sides + centered alignment, matching the real file -- previously
    skipped for Render-free-tier performance reasons that no longer apply
    now that this report runs via SSH offload to the office server.
  - Table tab is carried forward with its FORMULAS intact (not the cached
    values Excel last saved), so it keeps working exactly like it does when
    Khagash edits it by hand.

INDEPENDENTLY RE-VERIFIED (not just taking the provided script's word for it)
against the real files (Battery_Disconnection_Master_1_July_2026.xlsx as
"previous", Daily Battery Disconnected Consolidated Report 12th July 2026.xlsx
as the new day, MTR pull from 2026-07-13T17:10, compared against the real
Battery_Disconnection_Master_13_July_2026.xlsx):
  - Consolidated Shipment No. rows for 7/12: 80/80 exact, identical shipment
    sets both directions.
  - STO/NON STO, shareable link: 0 diffs / 80 rows.
  - Test, Onward Status: 32 diffs / 80 rows each -- CORRECTING the provided
    doc's explanation here: it claimed every diff was "our script reporting
    a MORE advanced state" (later MTR pull). Re-checking the actual diff
    VALUES directly disproves that as a clean explanation -- both directions
    occur (e.g. one row: ours = Pending, real = Stamp Verified, i.e. real is
    MORE advanced; another: ours = AI Complete, real = Pending, ours is more
    advanced; Test column diffs uniformly went the OPPOSITE direction from
    the doc's claim). This is most consistent with MTR being live,
    non-monotonic tracking data -- status fields aren't guaranteed to only
    move forward between arbitrary pulls -- rather than a script defect, but
    it should be described as "snapshot-sensitive, not fully understood",
    not as a confirmed one-directional timing artifact.
  - The join/filter/dedup pipeline itself (the part that actually matters
    for correctness) IS solid: exact row counts and shipment-set matches.

Still not automated (deliberately, per the reference doc): actual offline/
online determination is a manual human judgment call (Mail Status / vehicle
test columns), and actual mail-sending is out of scope -- this only builds
the Master candidate list.

I/O differences from the original script:
  - Inputs: 3 uploaded files (previous Master .xlsx, today's raw consolidated
    report .xlsx, latest MTR export) instead of CLI paths.
  - Output: streamed back in the same request instead of written to a local
    --output path.
  - Report Date is a single web-form date field (this report is inherently a
    single-day increment, same reasoning as Reports 3 & 5's date field).
"""

import logging
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from reports.errors import ReportProcessingError, describe_column_mismatch
from reports.registry import InputSlot, ReportMeta, register

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONFIG — column names kept exact to match the real workbook / source files
# ---------------------------------------------------------------------------

RAW_CONSOLIDATED_COLS = [
    "Vehicle No.", "Transporter Name", "Destination", "Event Lat.", "Event Long.",
    "Ship to Name", "Shipment No.", "Location Area", "Location Date and Time",
    "Invoice Date and Time", "Plant Name", "Status", "Unloading Point",
    "Ship to code", "Invoice No.", "Distribution Channel",
]

CSN_TAB_COLS = [
    "Mail Status", "vehicle test", "Date", "Vehicle No.", "Transporter", "Destination",
    "Event Lat.", "Event Long.", "Ship To Name", "Shipment No.", "Location Area",
    "Location Date and Time", "Invoice Date and Time", "Plant Name", "Status",
    "Unloading Point", "Ship to code", "Invoice No.", "Distribution Channel",
    "STO/NON STO", "Onward Status", "shareable link", "Test",
]

MASTER_TAB_COLS = [
    # NOTE: 4 of these headers use a non-breaking space (\xa0), not a regular
    # space -- confirmed by reading the real workbook's actual header cells.
    # Getting this wrong (regular space) silently creates duplicate columns
    # once real data lands in them, since pandas/Excel treat them as
    # different column names.
    "Shipment No.", "Invoice Number", "Plant Name", "Truck\xa0No", "Transporter\xa0Name",
    "Unloading point(ship-to-party)", "Billing date", "Tracking Link", "Sold To Code",
    "Name(Sold-To Party)", "Distribution\xa0Channel", "Region(Ship\xa0 To Party)",
    "District(Ship-To Party)", "QTY (MT)", "Freight loss", "Last Date", "Area",
]

FREIGHT_LOSS = 5000

# The real MTR tab is NOT simply the raw export's columns as-is -- it's the
# raw export with these 12 columns dropped (same list Reports 1/2/3 already
# use for the same reason). Confirmed empirically: raw export minus these
# 12 columns, in the raw export's own remaining column order, matches the
# real reference file's MTR tab both in column SET and ORDER exactly (121
# columns either way). Without this, "Shipment Number" and every other
# hardcoded-letter-referenced column (e.g. MTR!CI:CI, MTR!Y:Y) sits 1-12
# positions off from where the real workbook's formulas expect them,
# breaking every XLOOKUP that joins against MTR -- this was the actual
# cause of #N/A / 0 results seen in testing, not stale/dropped-off MTR data
# as initially assumed.
MTR_DROP_COLUMNS = [
    "Probable Unloading Count", "Probable Unloading Detention",
    "1 Km Geofence Start Time", "1 Km Geofence End Time", "1 Km Geofence Detention",
    "20 Km Geofence Start Time", "20 Km Geofence End Time", "20 Km Geofence Detention",
    "40 Km Geofence Start Time", "40 Km Geofence End Time", "40 Km Geofence Detention",
    "Lap Sharable Link",
]


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def _load_mtr_lookup(mtr_path: Path):
    """Load MTR raw export. Returns (slim lookup df, MTR-tab df).

    The returned MTR-tab df's column ORDER must exactly match the real
    workbook's own MTR tab -- the real Master/Consolidated Shipment No.
    tabs use XLOOKUP formulas that hardcode MTR column LETTERS (e.g.
    "MTR!CI:CI"), not column names. That real layout is NOT simply "the
    raw export's columns as exported" (confirmed: a fresh raw export had
    "Shipment Number" at column 89, but the real workbook's MTR tab has it
    at column 87 -- the export's own column order isn't stable day to
    day). It's "the raw export with MTR_DROP_COLUMNS removed, keeping the
    remaining columns' relative order" -- confirmed this exactly reproduces
    the real file's MTR tab, both column set and order. Do NOT additionally
    reindex/reorder beyond that drop (an earlier version called
    `.set_index("Shipment Number").reset_index()` before writing it out,
    which moved that column to the front and shifted everything else one
    position right -- also wrong, for the same reason).
    """
    try:
        if mtr_path.suffix.lower() == ".csv":
            raw = pd.read_csv(mtr_path, dtype={"Shipment Number": "Int64"})
        else:
            raw = pd.read_excel(mtr_path, dtype={"Shipment Number": "Int64"})
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read MTR Raw file '{mtr_path.name}': {exc}") from exc

    lookup = raw.drop_duplicates(subset=["Shipment Number"], keep="first").set_index("Shipment Number")
    keep_cols = ["Product Name", "40 Km Geofence Start Time"]

    mtr_tab_df = raw.drop(columns=[c for c in MTR_DROP_COLUMNS if c in raw.columns])
    return lookup[keep_cols], mtr_tab_df


def _load_new_consolidated_report(path: Path) -> pd.DataFrame:
    """Load today's raw Daily Battery Disconnected Consolidated Report (Sheet1)."""
    try:
        df = pd.read_excel(path, sheet_name="Sheet1")
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read '{path.name}': {exc}") from exc
    mismatch = describe_column_mismatch(df.columns, RAW_CONSOLIDATED_COLS, path.name)
    if mismatch:
        raise ReportProcessingError(
            f"{mismatch} Check you uploaded the Daily Battery Disconnected Consolidated Report."
        )
    return df[RAW_CONSOLIDATED_COLS].copy()


def _load_prev_master_tabs(path: Path):
    """Load the 3 tabs we carry forward from the previous Master workbook."""
    try:
        consolidated = pd.read_excel(path, sheet_name="Consolidated")
        csn = pd.read_excel(path, sheet_name="Consolidated Shipment No.")
        master = pd.read_excel(path, sheet_name="Master")
    except Exception as exc:
        raise ReportProcessingError(
            f"Couldn't read '{path.name}': {exc}. Check you uploaded the previous day's "
            "Battery_Disconnection_Master_<date>.xlsx (needs Consolidated, Consolidated "
            "Shipment No., Master tabs)."
        ) from exc
    return consolidated, csn, master


def _load_table_tab_rows(path: Path):
    """Read the 'Table' tab's raw cell contents -- FORMULAS included, not
    just their cached values, AND each cell's number_format -- using
    read_only mode so we never load the workbook's other, much larger
    sheets (Consolidated, MTR) into memory as full rich cell objects.

    Returns a list of rows, each a list of (value, number_format) tuples.

    data_only=False (not True) is deliberate: the real Table tab's rows are
    almost entirely XLOOKUP formulas referencing the Master tab (row 1's
    Shipment No. is the only literal value); reading with data_only=True
    would silently flatten them to whatever value happened to be cached the
    last time the source file was saved in Excel, permanently destroying
    the formulas on the very first run through this pipeline. Confirmed
    read_only=True still returns formula text (as a plain string) rather
    than a computed value when data_only=False.

    number_format matters here too: 2 of these rows (Billing date, Last
    Date) are date-valued XLOOKUPs formatted as "m/d/yy h:mm" in the real
    file -- reading only cell.value (via values_only=True) drops that,
    which otherwise makes Excel display the recalculated result as a raw
    date serial number (e.g. 46212.75) instead of a date/time once the
    formula re-evaluates against this run's freshly written Master tab.
    Confirmed read_only mode still exposes number_format per cell.

    openpyxl.load_workbook(path) (full/writable mode) on a growing multi-tab
    workbook measured 1.6GB peak memory and 188s locally -- far past
    Render's free-tier limits, and the actual cause of the 502s seen live.
    Read-only mode + only touching the one small tab we need avoids that.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read '{path.name}': {exc}") from exc
    if "Table" not in wb.sheetnames:
        return []
    ws = wb["Table"]
    rows = [[(cell.value, cell.number_format) for cell in row] for row in ws.iter_rows()]
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Pipeline steps
# ---------------------------------------------------------------------------

def build_new_consolidated_rows(new_raw_df: pd.DataFrame, report_date: pd.Timestamp) -> pd.DataFrame:
    """Tag today's raw rows with the report date, unfiltered (for 'Consolidated' tab).

    Renames to the same "Transporter"/"Ship To Name" headers the historical
    Consolidated tab (and the CSN tab, via build_new_csn_rows) already use --
    a prior version skipped this rename here, so freshly appended rows
    landed under "Transporter Name"/"Ship to Name" instead, splitting what
    should be one column into two once concatenated with history (confirmed
    by comparing our output against the real reference file: every new row
    had the renamed columns null and the raw-named columns populated, the
    reverse of every historical row).
    """
    df = new_raw_df.copy()
    df = df.rename(columns={"Transporter Name": "Transporter", "Ship to Name": "Ship To Name"})
    df.insert(0, "Date", report_date)
    return df


def build_new_csn_rows(new_raw_df: pd.DataFrame, report_date: pd.Timestamp, mtr_lookup: pd.DataFrame):
    """Clinker-filter + per-day dedupe + derive columns (for 'Consolidated Shipment No.').

    STO/NON STO, Onward Status, and shareable link are NOT computed here --
    in the real workbook these are live Excel formulas (an IFS lookup and
    two XLOOKUPs against the MTR tab), not static values, so they're left
    as placeholder blanks in this DataFrame and written as actual formula
    cells at output time instead (see FORMULA_SPECS / _write_sheet below).
    Computing static equivalents here would only approximate what the
    formulas do and can silently drift (confirmed: the real "Onward Status"
    formula actually reads MTR's "Stamp Status" column despite its own
    label, which a plain pandas .map() using the "Onward Status" MTR column
    would never replicate exactly).
    """
    df = new_raw_df.copy()

    # Join Product Name (only used here to filter Clinker; not stored in CSN tab).
    df = df.merge(
        mtr_lookup[["Product Name"]], left_on="Shipment No.", right_index=True, how="left"
    )
    before = len(df)
    df = df[~df["Product Name"].astype(str).str.lower().eq("clinker")]
    clinker_dropped = before - len(df)

    # Per-day dedupe, keep first occurrence in source file order.
    before2 = len(df)
    df = df.drop_duplicates(subset=["Shipment No."], keep="first")
    dupes_dropped = before2 - len(df)

    df.insert(0, "vehicle test", None)
    df.insert(0, "Mail Status", None)
    df.insert(2, "Date", report_date)

    df = df.rename(columns={
        "Transporter Name": "Transporter",
        "Ship to Name": "Ship To Name",
    })

    df["STO/NON STO"] = None  # written as a formula at output time
    df["Onward Status"] = None  # written as a formula at output time
    df["shareable link"] = None  # written as a formula at output time

    geofence = df["Shipment No."].map(mtr_lookup["40 Km Geofence Start Time"])
    found_in_mtr = df["Shipment No."].isin(mtr_lookup.index)
    entered_40km = geofence.notna() & (geofence.astype(str).str.strip() != "")
    df["Test"] = None
    df.loc[found_in_mtr & entered_40km, "Test"] = "Enter"
    df.loc[found_in_mtr & ~entered_40km, "Test"] = "Not Enter"
    # shipments not found in MTR: Test stays blank ("Unknown") rather than guessed

    df = df.drop(columns=["Product Name"])
    df = df[CSN_TAB_COLS]

    stats = {"clinker_dropped": clinker_dropped, "same_day_dupes_dropped": dupes_dropped}
    return df, stats


def build_master_additions(full_csn_df: pd.DataFrame, prev_master_df: pd.DataFrame) -> pd.DataFrame:
    """Find shipments newly qualifying for Master (vehicle test starts with 'offline').

    Only 'Shipment No.' and 'Freight loss' are real values here -- every
    other Master column is an XLOOKUP formula in the real workbook (against
    Consolidated Shipment No. or MTR), so they're left as placeholder
    blanks and written as formulas at output time instead (see
    FORMULA_SPECS / _write_sheet below), the same reasoning as
    build_new_csn_rows above.
    """
    df = full_csn_df.copy()
    df["vehicle test"] = df["vehicle test"].fillna("")
    df["_is_offline"] = df["vehicle test"].str.strip().str.lower().str.startswith("offline")

    offline_ships = set(df.loc[df["_is_offline"], "Shipment No."].unique())
    already_in_master = set(prev_master_df["Shipment No."].unique())

    # Exclude shipments whose most recent Mail Status is "Waived OFF"
    df_sorted = df.sort_values("Date")
    latest_status = df_sorted.groupby("Shipment No.")["Mail Status"].last()
    waived = set(latest_status[latest_status.astype(str).str.strip().str.lower() == "waived off"].index)

    new_ships = sorted(offline_ships - already_in_master - waived)
    if not new_ships:
        return pd.DataFrame(columns=MASTER_TAB_COLS)

    out = pd.DataFrame({"Shipment No.": new_ships})
    for col in MASTER_TAB_COLS:
        if col == "Shipment No.":
            continue
        out[col] = FREIGHT_LOSS if col == "Freight loss" else None
    return out[MASTER_TAB_COLS]


# ---------------------------------------------------------------------------
# Writer -- header fills/fonts/column widths extracted cell-by-cell from the
# real Battery_Disconnection_Master_13_July_2026.xlsx (data logic untouched;
# this section only affects display).
# ---------------------------------------------------------------------------

from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from reports._report6_col_widths import MTR_COL_WIDTHS

APTOS_10 = "Aptos Narrow"

# Per-sheet header style: (fill color, font name, font size).
SHEET_STYLES = {
    "Consolidated": {"fill": "FFC000", "font": APTOS_10, "size": 10},
    "Consolidated Shipment No.": {"fill": "8FAADC", "font": APTOS_10, "size": 10},
    "Master": {"fill": "FFFF00", "font": APTOS_10, "size": 10},
    "MTR": {"fill": "A9D18E", "font": "Calibri", "size": 11},
}

# Every data cell (not just headers) on Consolidated/CSN/Master uses this
# exact font + thin border on all 4 sides + centered alignment in the real
# workbook (confirmed cell-by-cell on the reference file). MTR's data cells
# use openpyxl's own defaults (Calibri 11, no border) so it needs no entry
# here. Shared Font/Border/Alignment instances (not one per cell) keep the
# O(n) styling loop below cheap even across tens of thousands of rows --
# this report now runs on the office server via SSH offload (real CPU/RAM),
# not Render's free tier, so this per-cell cost is no longer the risk it
# used to be when this sheet only got header styling.
DATA_FONT = Font(name=APTOS_10, size=10, bold=False)
_THIN = Side(style="thin")
DATA_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
SHEETS_WITH_FULL_DATA_STYLE = {"Consolidated", "Consolidated Shipment No.", "Master"}

# These columns hold live Excel formulas in the real workbook, not static
# values -- see build_new_csn_rows / build_master_additions above for why
# they're computed here instead of in pandas. Each function takes the
# 1-based Excel row number and returns either a formula string or an
# ArrayFormula (for the one column that's array-entered in the real file).
# Column letters (S, J, MTR!CI, MTR!Y, etc.) are hardcoded to exactly match
# the real workbook's own formulas, which assume the same fixed column
# layout this report also uses (CSN_TAB_COLS / MASTER_TAB_COLS order, and
# the MTR tab's untouched original export column order -- see
# _load_mtr_lookup's docstring).
FORMULA_SPECS = {
    "Consolidated Shipment No.": {
        "STO/NON STO": lambda r: ArrayFormula(
            f"T{r}",
            f'=_xlfn.IFS(S{r}=20,"DEALER TRADE",S{r}=10,"DIR PARTY(NON TRADE)",S{r}=30,"STOCK TRANSFER")',
        ),
        # NOTE: despite its "Onward Status" header, the real file's formula
        # reads MTR's "Stamp Status" column (MTR!Y:Y), not MTR's own
        # "Onward Status" column -- confirmed both columns exist and differ
        # in the raw MTR export. Replicated exactly as-is, not "corrected".
        "Onward Status": lambda r: f"=_xlfn.XLOOKUP(J{r},MTR!CI:CI,MTR!Y:Y,0)",
        "shareable link": lambda r: f"=_xlfn.XLOOKUP(J{r},MTR!CI:CI,MTR!BE:BE,0)",
    },
    "Master": {
        "Invoice Number": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!R:R)",
        "Plant Name": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!N:N)",
        "Truck\xa0No": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!D:D)",
        "Transporter\xa0Name": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!E:E)",
        "Unloading point(ship-to-party)": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!P:P)",
        "Billing date": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!M:M)",
        "Tracking Link": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!V:V)",
        "Sold To Code": lambda r: f"=_xlfn.XLOOKUP(A{r},MTR!CI:CI,MTR!AI:AI)",
        "Name(Sold-To Party)": lambda r: f"=_xlfn.XLOOKUP(A{r},MTR!CI:CI,MTR!AJ:AJ)",
        "Distribution\xa0Channel": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!S:S)",
        "Region(Ship\xa0 To Party)": lambda r: f"=_xlfn.XLOOKUP(A{r},MTR!CI:CI,MTR!CE:CE)",
        "District(Ship-To Party)": lambda r: f"=_xlfn.XLOOKUP(A{r},MTR!CI:CI,MTR!AT:AT)",
        "QTY (MT)": lambda r: f"=_xlfn.XLOOKUP(A{r},MTR!CI:CI,MTR!CV:CV)",
        "Last Date": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!L:L)",
        "Area": lambda r: f"=_xlfn.XLOOKUP(A{r},'Consolidated Shipment No.'!J:J,'Consolidated Shipment No.'!K:K)",
    },
}

# Despite the name, this is really "column name -> number_format override" --
# mostly dates, plus one cosmetic exception below.
DATE_COLUMN_FORMATS = {
    "Date": "mm-dd-yy",
    "Location Date and Time": "m/d/yy h:mm",
    "Invoice Date and Time": "m/d/yy h:mm",
    "Billing date": "mm-dd-yy",
    # Forces text display so long digit strings never render in scientific
    # notation -- matches what Khagash is used to seeing, doesn't affect
    # correctness. No pd.to_datetime() side effect here (unlike Report 3's
    # equivalent dict) since this report doesn't pre-parse date_col_idxs
    # entries before writing -- only number_format gets applied.
    "Transporter Number": "@",
}

CONSOLIDATED_COL_WIDTHS = {
    "Date": 10.11, "Vehicle No.": 12.33, "Transporter": 40.78, "Destination": 58.55,
    "Event Lat.": 12.0, "Ship To Name": 71.78, "Shipment No.": 11.66,
    "Location Area": 48.44, "Location Date and Time": 19.33, "Invoice Date and Time": 18.44,
    "Plant Name": 46.66, "Status": 17.44, "Unloading Point": 22.33, "Ship to code": 11.0,
    "Distribution Channel": 17.44,
}
CSN_COL_WIDTHS = {
    "Mail Status": 9.89, "vehicle test": 68.22, "Date": 10.11, "Vehicle No.": 12.33,
    "Transporter": 41.0, "Destination": 58.55, "Event Lat.": 12.0, "Ship To Name": 71.78,
    "Shipment No.": 11.66, "Location Area": 48.44, "Location Date and Time": 19.33,
    "Invoice Date and Time": 18.44, "Plant Name": 46.66, "Status": 17.44,
    "Unloading Point": 22.33, "Ship to code": 11.0, "Distribution Channel": 17.44,
    "STO/NON STO": 19.22, "Onward Status": 12.11, "shareable link": 72.33, "Test": 8.11,
}
MASTER_COL_WIDTHS = {
    "Shipment No.": 11.89, "Invoice Number": 13.78, "Plant Name": 12.22,
    "Transporter\xa0Name": 41.55, "Unloading point(ship-to-party)": 25.11,
    "Billing date": 10.33, "Tracking Link": 74.22, "Sold To Code": 11.22,
    "Name(Sold-To Party)": 51.33, "Distribution\xa0Channel": 17.44,
    "Region(Ship\xa0 To Party)": 18.55, "District(Ship-To Party)": 18.44,
    "QTY (MT)": 8.44, "Freight loss": 10.11, "Last Date": 12.11, "Area": 39.89,
}
DEFAULT_COL_WIDTH = 13.0

COL_WIDTHS_BY_SHEET = {
    "Consolidated": CONSOLIDATED_COL_WIDTHS,
    "Consolidated Shipment No.": CSN_COL_WIDTHS,
    "Master": MASTER_COL_WIDTHS,
    "MTR": MTR_COL_WIDTHS,
}


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame):
    """Plain (non write_only) Workbook + df.itertuples() bulk append --
    matches the fast pattern used in Reports 2 & 3 (write_only mode was
    tried and reverted, see git history: slower in practice here, not worth
    its memory saving on this specific workload). Header fill/font, column
    widths, date number-formats, per-cell data styling, and formula columns
    are all matched to the real file below; no freeze panes -- the real
    file doesn't use them either, despite the original script adding
    freeze_panes="A2" on every tab.

    Full per-cell data-row styling (font/border/alignment) on Consolidated/
    CSN/Master, and writing actual formulas for FORMULA_SPECS columns, used
    to be considered too expensive here (an O(n) styling loop across
    ~1.45M cells was the exact cost that made this report's write step
    marginal on Render's free tier). That's no longer the constraint this
    report runs under -- it now executes on the office server via SSH
    offload (real CPU/RAM), not Render, so the per-cell cost is acceptable
    and full-fidelity output (matching the real workbook cell-for-cell,
    formulas included) takes priority.
    """
    style = SHEET_STYLES.get(sheet_name, {"fill": None, "font": "Calibri", "size": 11})
    header_font = Font(name=style["font"], size=style["size"], bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor=style["fill"]) if style["fill"] else None
    col_widths = COL_WIDTHS_BY_SHEET.get(sheet_name, {})
    apply_full_data_style = sheet_name in SHEETS_WITH_FULL_DATA_STYLE
    formula_specs = FORMULA_SPECS.get(sheet_name, {})

    ws = wb.create_sheet(sheet_name)
    columns = list(df.columns)
    ws.append(columns)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx)
        cell.font = header_font
        if header_fill:
            cell.fill = header_fill
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(str(name).strip(), DEFAULT_COL_WIDTH)

    date_col_idxs = {i: DATE_COLUMN_FORMATS[name] for i, name in enumerate(columns) if name in DATE_COLUMN_FORMATS}
    formula_col_idxs = {i: formula_specs[name] for i, name in enumerate(columns) if name in formula_specs}

    row_idx = 1
    for row in df.itertuples(index=False):
        row_idx += 1
        ws.append(row)
        if apply_full_data_style:
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.font = DATA_FONT
                cell.border = DATA_BORDER
                cell.alignment = DATA_ALIGNMENT
        for idx, fmt in date_col_idxs.items():
            ws.cell(row_idx, idx + 1).number_format = fmt
        for idx, formula_fn in formula_col_idxs.items():
            ws.cell(row_idx, idx + 1).value = formula_fn(row_idx)
    return ws


def _write_table_tab(wb, rows: list):
    """Table is a tiny (~17-row) scratch/lookup helper, so per-cell styling
    here is cheap and safe (unlike the large data sheets above). Matches
    the real file cell-by-cell (confirmed, not assumed):
      - Every cell (both columns) gets thin border on all 4 sides + centered
        alignment.
      - Column A: Calibri 11 bold, solid yellow fill, on every row.
      - Column B: Calibri 11 plain on rows 2-14, but Aptos Narrow 10 (not
        bold) on row 1 specifically, and Calibri 11 BOLD on rows 15-17
        (Freight loss/Last Date/Area) -- an odd manual quirk in the real
        file, replicated exactly rather than "corrected" since matching
        the real file exactly is the point here.
      - number_format is preserved per-cell from the source file (see
        _load_table_tab_rows -- 2 rows are date-valued XLOOKUPs that need
        their "m/d/yy h:mm" format preserved, or Excel shows the
        recalculated result as a raw date serial number instead).

    `rows` is a list of rows, each a list of (value, number_format) tuples,
    as produced by _load_table_tab_rows.
    """
    ws = wb.create_sheet("Table")
    label_font = Font(name="Calibri", size=11, bold=True)
    label_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    value_font_row1 = Font(name=APTOS_10, size=10, bold=False)
    value_font_plain = Font(name="Calibri", size=11, bold=False)
    value_font_bold = Font(name="Calibri", size=11, bold=True)

    for row_idx, row in enumerate(rows, start=1):
        ws.append([value for value, _ in row])
        for col_idx, (_, number_format) in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx)
            # A cell openpyxl's read-only mode never actually touched (e.g.
            # a genuinely blank row) reports number_format as None rather
            # than "General" -- guard against that rather than writing an
            # invalid None into the output, which openpyxl refuses to save.
            cell.number_format = number_format or "General"
            cell.border = DATA_BORDER
            cell.alignment = DATA_ALIGNMENT
        label_cell = ws.cell(row_idx, 1)
        label_cell.font = label_font
        label_cell.fill = label_fill
        value_cell = ws.cell(row_idx, 2)
        if row_idx == 1:
            value_cell.font = value_font_row1
        elif row_idx >= 15:
            value_cell.font = value_font_bold
        else:
            value_cell.font = value_font_plain
    ws.column_dimensions["A"].width = 27.33
    ws.column_dimensions["B"].width = 75.11
    return ws


# ---------------------------------------------------------------------------
# Entry point used by the generic upload -> process -> download route
# ---------------------------------------------------------------------------

def process(input_files: dict, dates: dict, output_dir: Path) -> Path:
    report_date_str = dates["report_date"]
    report_date = pd.Timestamp(report_date_str)
    log.info("Processing Battery Disconnection Mail Creation for %s", report_date_str)

    mtr_lookup, mtr_full = _load_mtr_lookup(input_files["mtr_raw"])
    new_raw = _load_new_consolidated_report(input_files["new_consolidated"])
    prev_consolidated, prev_csn, prev_master = _load_prev_master_tabs(input_files["prev_master"])
    table_rows = _load_table_tab_rows(input_files["prev_master"])

    try:
        new_consolidated_rows = build_new_consolidated_rows(new_raw, report_date)
        full_consolidated = pd.concat([prev_consolidated, new_consolidated_rows], ignore_index=True)

        new_csn_rows, stats = build_new_csn_rows(new_raw, report_date, mtr_lookup)
        full_csn = pd.concat([prev_csn, new_csn_rows], ignore_index=True)

        new_master_rows = build_master_additions(full_csn, prev_master)
        full_master = pd.concat([prev_master, new_master_rows], ignore_index=True)
    except KeyError as exc:
        raise ReportProcessingError(
            f"Expected column {exc} not found. Check each file was uploaded to the correct "
            "slot (Previous Master / Today's Consolidated Report / MTR Raw)."
        ) from exc

    review_queue = new_csn_rows[
        (new_csn_rows["Test"] == "Not Enter") & (new_csn_rows["vehicle test"].isna())
    ]
    log.info(
        "Clinker dropped: %d | Same-day dupes dropped: %d | New Master rows: %d | "
        "Manual review queue: %d",
        stats["clinker_dropped"], stats["same_day_dupes_dropped"],
        len(new_master_rows), len(review_queue),
    )

    # Build a FRESH workbook rather than mutating the previous file's loaded
    # object -- loading the previous file in full/writable mode (just to
    # preserve its small 'Table' tab) pulled its huge Consolidated/MTR tabs
    # into memory as rich cell objects too -- measured 1.6GB peak / 188s
    # locally. Fixed by reading Table's values separately via read_only mode
    # (see _load_table_tab_rows) and starting fresh here. See _write_sheet's
    # docstring for why this uses a plain Workbook rather than write_only
    # mode despite the latter's much lower memory use.
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Consolidated", full_consolidated)
    _write_sheet(wb, "Consolidated Shipment No.", full_csn)
    _write_table_tab(wb, table_rows)
    _write_sheet(wb, "Master", full_master)
    # No reset_index()/reorder here -- mtr_full already preserves the raw
    # export's original column order, which the XLOOKUP formulas above
    # (hardcoded to column letters like MTR!CI:CI) depend on staying put.
    _write_sheet(wb, "MTR", mtr_full)

    output_path = output_dir / f"Battery_Disconnection_Master_{report_date_str}.xlsx"
    wb.save(output_path)
    return output_path


def process_dispatch(input_files: dict, dates: dict, output_dir: Path) -> Path:
    """Entry point wired into the registry below. Offloads to the office
    server over SSH when OFFICE_SERVER_HOST is configured (see
    core/ssh_worker.py + office_server_worker.py at the repo root) -- this
    report's write step is the one that has repeatedly OOM'd/timed out on
    Render's free tier (see notes above). With no office server configured
    (the default), falls straight through to the same `process()` above
    unchanged, so local dev/testing behaves exactly as it always has.
    """
    from core.ssh_worker import is_configured, run_remote

    if is_configured():
        return run_remote("6", input_files, dates, output_dir)
    return process(input_files, dates, output_dir)


# ---------------------------------------------------------------------------
# RE-ENABLED for another live test (per explicit instruction): the previous
# 3 production attempts 502'd (467MB/76s and 96MB/174s locally, both failed
# in the 53-127s range on Render). Since then: switched to itertuples-based
# writing (measured 42.6s/41.7s locally, down from the original 188s), and
# this pass added real-file formatting with no additional per-cell cost
# (confirmed no timing regression). Genuinely lower risk than the last
# attempt, but still unproven on Render's actual CPU/memory -- if this
# still 502s, the earlier stub-with-notes state is the fallback (see git
# history for that version).
# ---------------------------------------------------------------------------

register(
    ReportMeta(
        id="6",
        name="Battery Disconnection Mail Creation",
        input_slots=[
            InputSlot(
                key="prev_master",
                label="Previous Day's Master Workbook",
                accept=".xlsx",
                hint="Battery_Disconnection_Master_<prev_date>.xlsx",
            ),
            InputSlot(
                key="new_consolidated",
                label="Today's Daily Battery Disconnected Consolidated Report",
                accept=".xlsx",
                hint="Daily Battery Disconnected Consolidated Report <date>.xlsx",
            ),
            InputSlot(key="mtr_raw", label="MTR Raw", accept=".csv,.xlsx", hint="mtr - <timestamp>.csv"),
        ],
        output_pattern="Battery_Disconnection_Master_<date>.xlsx (5 tabs: Consolidated, Consolidated "
                      "Shipment No., Table, Master, MTR)",
        # implemented=False is COSMETIC ONLY here -- process_fn is still the
        # real, working `process` function below, not the not_implemented
        # stub. main.py's /generate route never checks `implemented`; it
        # only drives the "not yet implemented" badge/text on the report
        # page. Per explicit instruction: show the stub label (since this
        # report is still unproven on Render's free tier and shouldn't be
        # advertised as ready), but if someone uploads and clicks Generate
        # anyway, it should still produce a correct, fully-working output --
        # which this achieves, since the real process_fn is still wired in.
        process_fn=process_dispatch,
        implemented=False,
        date_mode="single",
        notes=(
            "Stateful report -- needs the PREVIOUS day's Master workbook as an input each run "
            "(carries forward history + any manual Mail Status / vehicle test tags). Master tab = "
            "shipments whose 'vehicle test' starts with 'offline', append-only, excluding 'Waived "
            "OFF'. Formatting matches the real master file exactly. Previously 502'd on Render's "
            "free tier 3 times before write-speed optimizations (188s -> ~42s locally) -- being "
            "retested live; if it 502s again this needs a hosting upgrade, not more code tuning."
        ),
    )
)
