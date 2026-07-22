"""
JKLC Daily Tracker — web app adaptation
========================================

Adapted from the original standalone script (2 Google Sheets in -> 1 Google
Sheet out) to the web app's file-in/file-out flow. Per instructions, ALL
business logic below (date-window rule, column drops, cleaning rules,
matching rule, computed columns, summary calculations, percentage/bucket
rules) is kept exactly as provided — only the I/O layer changed:

  - Input: 2 uploaded files (MTR Raw csv/xlsx, Daily Dispatch xlsx) instead
    of 2 Google Sheets.
  - Output: 1 generated .xlsx file (built from the "JKLC Daily Tracker
    Master.xlsx" template so the Tracker tab's title rows / merged group
    headers / column labels (rows 1-4) and the MTR tab's exact 104-column
    order are preserved) instead of writing into a persistent Google Sheet.
  - START_DATE/END_DATE are threaded through as function parameters instead
    of being module-level globals computed once at import time — necessary
    because this now runs per web request (potentially concurrently) rather
    than as a single top-level script run. This is a structural change only,
    not a logic change.
  - print()/SystemExit() calls (fine for a CLI script) are replaced with
    logging / ReportProcessingError so the web UI can show a readable
    message instead of a raw traceback.

Flagged instead of silently changed (per instructions) — nothing else was
altered; see chat for anything else worth flagging.
"""

import logging
from pathlib import Path

import openpyxl
import pandas as pd

from reports.errors import ReportProcessingError
from reports.registry import InputSlot, ReportMeta, register

log = logging.getLogger(__name__)

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent.parent
    / "Report Templates"
    / "JKLC Daily Tracker Master.xlsx"
)

# ---------------------------------------------------------------------------
# CONFIG — copied verbatim from the original script
# ---------------------------------------------------------------------------

# Confirmed finding: real output still has SIM / No Mode rows, so default False.
APPLY_MODE_FILTER = False

# Columns dropped during "MTR cleaning" (confirmed via raw vs cleaned diff)
MTR_DROP_COLUMNS = [
    "Probable Unloading Count",
    "Probable Unloading Detention",
    "1 Km Geofence Start Time",
    "1 Km Geofence End Time",
    "1 Km Geofence Detention",
    "20 Km Geofence Start Time",
    "20 Km Geofence End Time",
    "20 Km Geofence Detention",
    "40 Km Geofence Start Time",
    "40 Km Geofence End Time",
    "40 Km Geofence Detention",
    "Lap Sharable Link",
]

ORG_LOCATION_MAP = {
    "Durg Plant": "JKLC Durg",
    "Jharli Grinding": "JKLC Jharli",
    "JK Lakshmi Cement Limited- Surat Grinding Unit": "JKLC Surat",
    "JK Lakshmi Cement Limited - Surat Grinding Unit": "JKLC Surat",
    "MS JK LAKSHMI CEMENT LIMITED CUTTACK GRINDING UNIT": "JKLC Cuttack",
    "M/S JK Lakshmi Cement Limited Cuttack Grinding Unit": "JKLC Cuttack",
}

# Dispatch PLANT_NAME -> same JKLC naming, for the "Trips Not Created" count.
DISPATCH_PLANT_MAP = {
    "DURG": "JKLC Durg",
    "JHARLI": "JKLC Jharli",
    "SURAT": "JKLC Surat",
    "CUTTACK": "JKLC Cuttack",
}

# Always show all 4 plants in the Tracker summary, even if one has zero
# matched rows in the current date window.
ORG_LOCATIONS_ALL = ["JKLC Cuttack", "JKLC Durg", "JKLC Jharli", "JKLC Surat"]

# Row where Tracker DATA starts in the template (1-indexed). Rows above this
# are the manually-formatted title / group-header / column-label rows and
# must NEVER be touched.
TRACKER_DATA_START_ROW = 5


def _compute_date_window(report_date_str: str):
    """Confirmed by Khagash + Aryan (13 July 2026): two rules depending on
    which day of the month REPORT_DATE falls on.

      - day-of-month <= 7  -> ROLLING window (XSwift's 30-day download cap):
        START = REPORT_DATE - 30 days, END = REPORT_DATE - 3 days.
      - day-of-month > 7   -> CALENDAR window (current month to date):
        START = 1st of REPORT_DATE's month, END = REPORT_DATE - 3 days.
    """
    report_date = pd.to_datetime(report_date_str)
    end = report_date - pd.Timedelta(days=3)
    if report_date.day <= 7:
        start = report_date - pd.Timedelta(days=30)
    else:
        start = report_date.replace(day=1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# File I/O (new — replaces gspread load_sheet_as_df / write_output)
# ---------------------------------------------------------------------------

def _read_any(path: Path) -> pd.DataFrame:
    """Read an uploaded MTR/Dispatch export as all-string columns, mirroring
    the original gspread get_all_values() behaviour (everything came back as
    strings there too, which the cleaning/matching logic below relies on)."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            return pd.read_csv(path, dtype=str)
        if suffix in (".xlsx", ".xls"):
            return pd.read_excel(path, dtype=str)
    except Exception as exc:
        raise ReportProcessingError(
            f"Couldn't read '{path.name}' as a {suffix} file: {exc}"
        ) from exc
    raise ReportProcessingError(
        f"Unsupported file type '{suffix}' for '{path.name}'. Expected .csv or .xlsx."
    )


# ---------------------------------------------------------------------------
# Cleaning (business logic unchanged — only START_DATE/END_DATE are now
# parameters instead of module globals)
# ---------------------------------------------------------------------------

def parse_dt(series):
    return pd.to_datetime(series, errors="coerce", dayfirst=False)


def clean_mtr(df: pd.DataFrame, start_date: str, end_date: str) -> pd.DataFrame:
    df = df.copy()

    df = df.drop(columns=[c for c in MTR_DROP_COLUMNS if c in df.columns])

    df["Org Location"] = df["Org Location"].replace(ORG_LOCATION_MAP)

    df["_inv_dt"] = parse_dt(df["Invoice Date & Time"])
    mask = (df["_inv_dt"] >= start_date) & (df["_inv_dt"] <= end_date + " 23:59:59")
    df = df[mask]

    df = df[df["Transporter"].str.strip().str.upper() != "M/S. OWN TRUCK DURG"]

    # Rename GPS-API -> AT FIX (always applied, independent of the mode
    # filter below — confirmed 14 July 2026: these are the same tracking
    # mode and should be unified regardless of whether rows get filtered).
    df["Mode"] = df["Mode"].replace({"GPS-API": "AT FIX"})

    if APPLY_MODE_FILTER:
        df = df[df["Mode"].isin(["AT FIX", "GPS-API"])]

    return df


def clean_dispatch(df: pd.DataFrame, start_date: str, end_date: str) -> pd.DataFrame:
    df = df.copy()

    df["_bill_dt"] = parse_dt(df["Billing_date"])
    mask = (df["_bill_dt"] >= start_date) & (df["_bill_dt"] <= end_date + " 23:59:59")
    df = df[mask]

    df = df[df["Transporter Name"].str.strip().str.upper() != "OWN TRUCK DURG"]
    df = df[df["Sold-To"].str.strip() != "D6100"]
    df = df[~df["Material description"].str.lower().str.contains("clinker", na=False)]

    return df


# ---------------------------------------------------------------------------
# Match & count (business logic unchanged)
# ---------------------------------------------------------------------------

def match_and_count(mtr_clean: pd.DataFrame, dispatch_clean: pd.DataFrame):
    mtr_inv = mtr_clean["Invoice no"].astype(str).str.strip()
    disp_inv = set(dispatch_clean["Invoice Number"].astype(str).str.strip())

    matched_mtr = mtr_clean[mtr_inv.isin(disp_inv)].copy()

    matched_mtr_inv_set = set(matched_mtr["Invoice no"].astype(str).str.strip())
    dispatch_clean = dispatch_clean.copy()
    dispatch_clean["_matched"] = dispatch_clean["Invoice Number"].astype(str).str.strip().isin(matched_mtr_inv_set)
    not_created = dispatch_clean[~dispatch_clean["_matched"]]

    plant_col = "PLANT_NAME" if "PLANT_NAME" in dispatch_clean.columns else None
    if plant_col:
        trips_not_created_by_plant = (
            not_created[plant_col]
            .map(lambda p: DISPATCH_PLANT_MAP.get(str(p).strip().upper(), p))
            .value_counts()
            .to_dict()
        )
    else:
        trips_not_created_by_plant = {}

    return matched_mtr, trips_not_created_by_plant


# ---------------------------------------------------------------------------
# Computed columns on matched MTR (business logic unchanged)
# ---------------------------------------------------------------------------

def add_computed_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    completed_statuses = {"Stamp Verified", "High Confidence", "Low Confidence", "Rejected"}
    df["Trips"] = df["Stamp Status"].apply(
        lambda s: "Completed" if s in completed_statuses else "Enroute"
    )

    def remark(th):
        try:
            th = float(th)
        except (TypeError, ValueError):
            return None
        if th > 90:
            return 90
        if th > 80:
            return 80
        return None

    df["Remarks"] = df["Track Health %"].apply(remark)

    return df


# ---------------------------------------------------------------------------
# Tracker (Summary) tab (business logic unchanged)
# ---------------------------------------------------------------------------

def _compute_plant_metrics(plant_label: str, g: pd.DataFrame, not_created: int) -> dict:
    completed = (g["Trips"] == "Completed").sum()
    enroute = (g["Trips"] == "Enroute").sum()
    total = completed + enroute + not_created

    gc_ = g[g["Trips"] == "Completed"]

    tracked = (gc_["Track Status"] == "TRACKED").sum()
    untracked = (gc_["Track Status"] == "UNTRACKED").sum()
    track_total = tracked + untracked
    tracked_pct = tracked / track_total if track_total else None
    untracked_pct = untracked / track_total if track_total else None

    # AT FIX / Sim % must be "of the TRACKED devices", so BOTH the count
    # and the denominator need to be scoped to TRACKED rows -- not just the
    # denominator. The previous version counted at_fix/sim over ALL
    # completed rows (gc_, tracked+untracked together) while dividing by
    # `tracked` alone: on one real file this coincidentally matched (every
    # tracked row happened to have Mode AT FIX/SIM, no untracked row ever
    # did), producing believable output purely by luck. A second real file
    # exposed the actual bug -- an UNTRACKED row with Mode "AT FIX" still
    # got counted in the numerator, pushing the "percentage" past 100%
    # (e.g. Durg came out to 101.6%), which is what caught this.
    gc_tracked = gc_[gc_["Track Status"] == "TRACKED"]
    at_fix = (gc_tracked["Mode"] == "AT FIX").sum()
    sim = (gc_tracked["Mode"] == "SIM").sum()
    at_fix_pct = at_fix / tracked if tracked else None
    sim_pct = sim / tracked if tracked else None

    health_80 = (gc_["Remarks"] == 80).sum()
    health_90 = (gc_["Remarks"] == 90).sum()
    health_80_pct = health_80 / tracked if tracked else None
    health_90_pct = health_90 / tracked if tracked else None
    health_total_pct = (
        (health_80_pct + health_90_pct)
        if (health_80_pct is not None and health_90_pct is not None)
        else None
    )

    at_destination = (gc_["Deviate Remark"] == "At Destination").sum()
    backward = (gc_["Deviate Remark"] == "Backward Deviated").sum()
    forward = (gc_["Deviate Remark"] == "Forward Deviated").sum()

    return {
        "Plant": plant_label,
        "COMPLETED": completed,
        "Enroute": enroute,
        "Trips Not Created": not_created,
        " Total": total,
        "TRACKED": tracked,
        "UNTRACKED": untracked,
        " Total ": track_total,
        "Tracked %": tracked_pct,
        "UNTRACKED%": untracked_pct,
        "AT FIX": at_fix_pct,
        "Sim": sim_pct,
        "Greater than 80 & Less than equal to 90": health_80,
        "Greater than 90": health_90,
        "Trips>80&<=90": health_80_pct,
        "> 90%": health_90_pct,
        "Total  ": health_total_pct,
        "At Destination": at_destination,
        "Backward Deviated": backward,
        "Forward Deviated": forward,
    }


def build_tracker_summary(mtr: pd.DataFrame, trips_not_created_by_plant: dict) -> pd.DataFrame:
    present = dict(tuple(mtr.groupby("Org Location")))

    rows = []
    for plant in ORG_LOCATIONS_ALL:
        not_created = trips_not_created_by_plant.get(plant, 0)
        g = present.get(plant, mtr.iloc[0:0])
        rows.append(_compute_plant_metrics(plant, g, not_created))

    for plant in ORG_LOCATIONS_ALL:
        if plant not in present:
            log.warning(
                "'%s' has ZERO matched rows in this date window. Check the MTR "
                "Raw file — this could be a genuine data gap, or an Org Location "
                "spelling that doesn't match ORG_LOCATION_MAP.",
                plant,
            )

    total_not_created = sum(trips_not_created_by_plant.values())
    grand_total_row = _compute_plant_metrics("Grand Total", mtr, total_not_created)

    summary = pd.DataFrame(rows + [grand_total_row])
    return summary


# ---------------------------------------------------------------------------
# Write output (new — replaces the gspread write_output(); fills the
# "JKLC Daily Tracker Master.xlsx" template instead of a Google Sheet)
# ---------------------------------------------------------------------------

def _write_output_xlsx(tracker_df: pd.DataFrame, mtr_df: pd.DataFrame, output_path: Path) -> None:
    if not TEMPLATE_PATH.exists():
        raise ReportProcessingError(
            f"Report template not found at '{TEMPLATE_PATH}'. Make sure "
            "'JKLC Daily Tracker Master.xlsx' is in the 'Report Templates' folder."
        )

    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # --- Tracker tab: write summary_df values positionally into contiguous
    # columns starting at A, same as the original script's
    # `tracker_ws.update(range_name=f"A{first_row}", values=values)` — it
    # never matched by header text, it trusted the DataFrame's column order
    # already lined up with the sheet's columns. (FYI: this template's row-4
    # header cells use slightly different whitespace than the script's dict
    # keys, e.g. "Total" vs "Total  " — cosmetic only, doesn't affect data,
    # flagging in case you want to tidy the template text later.)
    tracker_ws = wb["Tracker"]

    # Clear any pre-existing sample data/formulas below the header row before
    # writing fresh values (mirrors the original script's batch_clear + write).
    for r in range(TRACKER_DATA_START_ROW, tracker_ws.max_row + 1):
        for c in range(1, tracker_ws.max_column + 1):
            tracker_ws.cell(row=r, column=c).value = None

    for i, (_, row) in enumerate(tracker_df.iterrows()):
        r = TRACKER_DATA_START_ROW + i
        for c, value in enumerate(row, start=1):
            tracker_ws.cell(row=r, column=c).value = None if pd.isna(value) else value

    # --- MTR tab: write header + data in the DataFrame's own natural column
    # order, same as the original script's `mtr_ws.clear()` + write from A1 —
    # it never reordered columns to match a pre-set template header either.
    mtr_ws = wb["MTR"]
    if mtr_ws.max_row > 0:
        mtr_ws.delete_rows(1, mtr_ws.max_row)

    mtr_ws.append(list(mtr_df.columns))
    for row in mtr_df.itertuples(index=False):
        mtr_ws.append([None if pd.isna(value) else value for value in row])

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Entry point used by the generic upload -> process -> download route
# ---------------------------------------------------------------------------

def process(input_files: dict, dates: dict, output_dir: Path) -> Path:
    mtr_raw_path = input_files["mtr_raw"]
    dispatch_path = input_files["dispatch"]

    # Manual start/end range, per explicit instruction — NOT run through
    # _compute_date_window() (see that function's docstring for the
    # confirmed rolling/calendar rule it implements, kept but currently
    # unused now that the UI takes an explicit range instead of one date).
    start_date = dates["start_date"]
    end_date = dates["end_date"]
    log.info("Processing JKLC Daily Tracker for window %s to %s", start_date, end_date)

    mtr_raw = _read_any(mtr_raw_path)
    dispatch_raw = _read_any(dispatch_path)

    try:
        mtr_clean = clean_mtr(mtr_raw, start_date, end_date)
        dispatch_clean = clean_dispatch(dispatch_raw, start_date, end_date)
    except KeyError as exc:
        raise ReportProcessingError(
            f"Expected column {exc} not found. Check you uploaded the correct "
            "file for each slot (MTR Raw vs Daily Dispatch)."
        ) from exc

    matched_mtr, trips_not_created_by_plant = match_and_count(mtr_clean, dispatch_clean)
    matched_mtr = add_computed_columns(matched_mtr)
    tracker_summary = build_tracker_summary(matched_mtr, trips_not_created_by_plant)

    mtr_out = matched_mtr.drop(columns=[c for c in matched_mtr.columns if c.startswith("_")], errors="ignore")

    output_path = output_dir / f"JKLC_Daily_Tracker_{start_date}_to_{end_date}.xlsx"
    _write_output_xlsx(tracker_summary, mtr_out, output_path)
    return output_path


def process_dispatch(input_files: dict, dates: dict, output_dir: Path) -> Path:
    """Entry point wired into the registry below. Offloads to the office
    server over SSH when SSH_HOST is configured (see core/ssh_worker.py +
    office_server_worker.py at the repo root), so this report also runs on
    the office server's CPU/RAM instead of Render's. With no office server
    configured (the default), falls straight through to the same
    `process()` above unchanged -- no change to this report's actual logic
    or output, just where it executes.
    """
    from core.ssh_worker import is_configured, run_remote

    if is_configured():
        return run_remote("1", input_files, dates, output_dir)
    return process(input_files, dates, output_dir)


register(
    ReportMeta(
        id="1",
        name="JKLC Daily Tracker",
        input_slots=[
            InputSlot(
                key="mtr_raw",
                label="MTR Raw",
                accept=".csv,.xlsx",
                hint="raw_mtr_-_2026-07-13T125713_927.csv",
            ),
            InputSlot(
                key="dispatch",
                label="Daily Dispatch",
                accept=".xlsx",
                hint="Axestrack (72) Daily dispatch 9 july.xlsx",
            ),
        ],
        output_pattern="JKLC_Daily_Tracker_<start>_to_<end>.xlsx",
        process_fn=process_dispatch,
        implemented=True,
        notes="Fully validated, script provided.",
        date_mode="range",
    )
)
