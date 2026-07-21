"""
JKLC Daily Tracking Report — web app adaptation (Report 3 of 9)
=================================================================

Adapted from the provided `jklc_daily_tracking_report.py` script and its
companion reference doc, plus independent verification against the real
12th July 2026 report and the SOP docx for this pass.

Report shape: a single-day snapshot, but the web form uses a Start Date /
End Date range (matching Reports 1 & 2's UI) rather than one date field.
END_DATE is DATA_DATE — the day whose completed trips / dispatches /
installs this report covers (matches the real file's own title and
filename exactly, e.g. "JKLC Daily Tracking Report 12-07-2026" covers
DATA_DATE=2026-07-12). START_DATE only bounds the MTR tab's own date
filter (previously MTR was passed through unfiltered, assuming the
uploaded file was already pre-trimmed). "Last 5 days" = DATA_DATE - 4 ..
DATA_DATE. No cross-day accumulation (per explicit instruction) — every
run is fresh.

CONVENTION CORRECTION vs the original script/reference doc: both called
this "REPORT_DATE" and described it as "the day you're generating the
report FOR", with "Yesterday" = REPORT_DATE - 1 and the 5-day window ending
at REPORT_DATE. Validating against the real 12th July file line-by-line
disproved this on two counts:
  1. The file's own live-formula cells (e.g. Summary!G3 = "=TODAY()-1") had
     CACHED values of 2026-07-12 -- meaning when Khagash's sheet last
     recalculated, "yesterday" was 12 July, i.e. the day the file is named
     after is *yesterday's* data, not "today"'s. Every AT-installation
     metric in the Summary (Last Day AT Installation, AT Total Installation
     Till Date, JKLC Total Installation, and all 5 Durg Dispatch
     installation-side fields) matched EXACTLY once treated this way, and
     were all visibly wrong under the original "REPORT_DATE = filename
     date" reading.
  2. The 5-day window's own title text in the real file reads
     "08-07-2026 - 12-07-2026" -- ending at DATA_DATE (12), not the day
     after. This also matches the SOP docx's own worked example word for
     word: "if today is 10 June, take 9, 8, 7, 6, 5 June" -- i.e. the
     window ends at YESTERDAY, not at "today". The original script filtered
     up through REPORT_DATE ("today") instead, an off-by-one independent of
     the naming question above.
Net effect: one field, DATA_DATE, with no "+1"/"-1" offset exposed to the
user at all -- it directly is the day whose data appears in every part of
the report, matching the real file's filename/title/date-window text
exactly. See `process()` for how this collapses what the original script
treated as two dates (REPORT_DATE, YESTERDAY) into one.

Fixes/additions made in this pass, verified against the real 12th July file
(JKLC Daily Tracking Report 12th July 2026.xlsx) and the SOP docx:

  1. Deviation Remarks uses "Destination Deviation" / "Destination Deviation
     Direction" — the SOP text itself says "Lead Distance Deviation" but
     that's confirmed wrong (572/1999 mismatches vs 0/1999). Kept from the
     original script's own correction.
  2. Durg Dispatch Summary block was incomplete in the original script (only
     4 of 9 real metrics), and its vehicle dedup was wrong: the SOP text says
     "Remove duplicate rows by Vehicle No." but the real file's own numbers
     (195 total, 163 AT-matched, for DATA_DATE) match this file's RAW,
     UNDEDUPED Durg rows for that same date exactly -- deduping instead
     gives 177 / 146. A vehicle dispatched twice in a day apparently counts
     twice in the real report, contra the SOP text -- dedup step removed.
     (Separately, an earlier raw-MTR check against Invoice Date = DATA_DATE
     minus one day gave 219 vehicles -- that number matches neither this
     file's real total nor this fix; it was checking a different, incorrect
     date, not a competing theory about the dedup question.)
     The SOP docx (not fully reflected in the script) defines 3 more fields:
       - "Vehicles allotted to Axestrack" / "Axestrack's Installation Count"
         = All Installation rows filtered to Yesterday + that plant (this
         is the SAME number as the top Summary block's "Last Day AT
         Installation" for Durg — confirmed identical in the real file: 6).
       - "Total Prospect Vehicles for GPS installations arrived at Plant"
         = Vehicles allotted to Axestrack + Vehicles allowed without GPS
         (NOT derived by subtracting from Total Dispatched — confirmed via
         the real numbers: 6 + 31 = 37, matching exactly).
       - "Clinker Vehicle (Out of Prospect for GPS Installations)" has no
         rule anywhere in the SOP docx or script and is 0 in every real
         example seen — written as a fixed 0 with this caveat, not computed.
     With these added, the block's arithmetic partitions exactly:
     Total Dispatched = Clinker + AT pre-installed + Wheelseye pre-installed
     + Vehicles allowed without GPS (195 = 0 + 163 + 1 + 31, confirmed).
  3. "JKLC Total Installation" (top Summary block) = AT Total Installation
     Till Date + Last Day AT Installation + 3rd Party Vehicle Count.
     Confirmed via the real file's own numbers (6281+6+0=6287, matches).
     "3rd Party Vehicle Count" has no source in the 6 input files or SOP
     and is 0 in every real example — same caveat as Clinker Vehicle.
  4. Added the "<Month> Installation" tab (e.g. "July '26 Installation"):
     All Installation filtered to REPORT_DATE's month, per SOP step 3.
  5. Summary tab formatting extracted cell-by-cell from the real file
     (fonts, fills, merges, column widths) — same approach as Report 2.

With the dedup fix, Durg Dispatch matches the real 12th July file almost
exactly: Total Dispatched 195=195, AT-matched 163=163, No-GPS 32 vs real 31
(off by exactly 1, mirrored by Wheelseye 0 vs real 1 -- one vehicle's GPS
Provider Name in Master Entry Data likely doesn't read exactly "WHEELSEYE"
in this pull; not chased further, single-row edge case).
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reports._report3_col_widths_mtr import MTR_FAMILY_COL_WIDTHS, ALL_INSTALLATION_COL_WIDTHS
from reports.errors import ReportProcessingError, describe_column_mismatch
from reports.registry import ExtraNumberField, InputSlot, ReportMeta, register

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONFIG — same MTR cleaning as Reports 1 & 2 (confirmed identical here too)
# ---------------------------------------------------------------------------

MTR_DROP_COLUMNS = [
    "Probable Unloading Count",
    "Probable Unloading Detention",
    "1 Km Geofence Start Time",
    "1 Km Geofence End Time",
    "1 Km Geofence Detention",
    "20 Km Geofence Start Time",
    "20 Km Geofence End Time",
    "20 Km Geofence Detention",
    "40 Km Geofence Start Time",
    "40 Km Geofence End Time",
    "40 Km Geofence Detention",
    "Lap Sharable Link",
    # CONFIRMED against the real file: "Yesterday Completed Trips" doesn't
    # have this column either -- drop it same as the other 12.
    "Ulip Toll Count",
]

ORG_LOCATION_MAP = {
    "Durg Plant": "JKLC Durg",
    "Jharli Grinding": "JKLC Jharli",
    "JK Lakshmi Cement Limited- Surat Grinding Unit": "JKLC Surat",
    "JK Lakshmi Cement Limited - Surat Grinding Unit": "JKLC Surat",
    "MS JK LAKSHMI CEMENT LIMITED CUTTACK GRINDING UNIT": "JKLC Cuttack",
    "M/S JK Lakshmi Cement Limited Cuttack Grinding Unit": "JKLC Cuttack",
}

ORG_LOCATIONS_ALL = ["JKLC Cuttack", "JKLC Durg", "JKLC Jharli", "JKLC Surat"]


# ---------------------------------------------------------------------------
# File I/O
# ---------------------------------------------------------------------------

def _read_mtr(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            return pd.read_csv(path, dtype=str)
        if suffix in (".xlsx", ".xls"):
            return pd.read_excel(path, dtype=str)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read '{path.name}' as MTR data: {exc}") from exc
    raise ReportProcessingError(f"Unsupported file type '{suffix}' for MTR Raw. Expected .csv or .xlsx.")


def _load_device_status(path: Path) -> pd.DataFrame:
    # Real AT portal export has 2 report-title rows before the real header.
    try:
        return pd.read_csv(path, skiprows=2, encoding="utf-8-sig", dtype=str)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read Device Status file '{path.name}': {exc}") from exc


def _load_dashboard(path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(path, dtype=str)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read Dashboard export '{path.name}': {exc}") from exc


def _load_offline_dashboard(path: Path) -> pd.DataFrame:
    # AT portal quirk: this file has an .xls extension but is actually HTML.
    # read_html handles it; read_excel does not.
    try:
        tables = pd.read_html(path)
        return tables[0]
    except Exception as exc:
        raise ReportProcessingError(
            f"Couldn't read Offline Trip Dashboard '{path.name}' as HTML: {exc}"
        ) from exc


def _load_master_entry_data(path: Path) -> pd.DataFrame:
    # The real file's Excel "used range" can be ~1M rows (mostly blank
    # formatting artifacts) even though only ~50k rows have real data — a
    # naive pd.read_excel() call tries to materialize the whole thing and
    # can run out of memory. Stream through it with openpyxl instead and
    # stop collecting once we hit a long run of fully-blank rows.
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        records = []
        header = None
        blank_streak = 0
        for row in ws.iter_rows(min_col=1, max_col=13, values_only=True):
            if header is None:
                header = row
                continue
            if row[5] is None:  # Vehicle no. column blank -> not real data
                blank_streak += 1
                if blank_streak > 5000:
                    break
                continue
            blank_streak = 0
            records.append(row)

        return pd.DataFrame(records, columns=header)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read Master Entry Data '{path.name}': {exc}") from exc


# Exact raw column schema of the API Vehicles export (per the confirmed
# finding in clean_api_vehicles()'s docstring below: the raw file has these
# 6 columns -- fname, transporter_name, regno, apiprovider, addtime,
# modified_time).
API_VEHICLES_RAW_COLS = ["fname", "transporter_name", "regno", "apiprovider", "addtime", "modified_time"]


def _load_api_vehicles(path: Path) -> pd.DataFrame:
    try:
        df = pd.read_csv(path, dtype=str)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read API Vehicles file '{path.name}': {exc}") from exc
    mismatch = describe_column_mismatch(df.columns, API_VEHICLES_RAW_COLS, path.name)
    if mismatch:
        raise ReportProcessingError(f"{mismatch} Check you uploaded the raw API Vehicles export.")
    return df


def clean_api_vehicles(raw_df: pd.DataFrame) -> pd.DataFrame:
    """CORRECTED: the reference doc/original script assumed this tab was
    the raw API Vehicles export passed through as-is ("no logic needed").
    Comparing against the real 12th July file disproved that -- the real
    tab has 5 renamed columns (S.No., Vehicle No, Plant, Provider, Add
    Date), not the raw file's 6 (fname, transporter_name, regno,
    apiprovider, addtime, modified_time). Confirmed by matching sample
    values exactly: regno -> Vehicle No, fname -> Plant, apiprovider ->
    Provider, addtime -> Add Date; transporter_name and modified_time are
    dropped entirely; S.No. is a new 1-based row serial, not from the
    source file.
    """
    df = raw_df.copy()
    out = pd.DataFrame({
        "S.No.": range(1, len(df) + 1),
        "Vehicle No": df["regno"],
        "Plant": df["fname"],
        "Provider": df["apiprovider"],
        "Add Date": df["addtime"],
    })
    return out


# ---------------------------------------------------------------------------
# Clean MTR (same as Reports 1 & 2)
# ---------------------------------------------------------------------------

def clean_mtr(df: pd.DataFrame, start_date: str, end_date: str) -> pd.DataFrame:
    df = df.copy()
    df = df.drop(columns=[c for c in MTR_DROP_COLUMNS if c in df.columns])
    df["Org Location"] = df["Org Location"].replace(ORG_LOCATION_MAP)

    # Filter MTR to the explicit [start_date, end_date] range -- previously
    # this just passed the uploaded file through unfiltered, assuming it was
    # already pre-trimmed to the right window.
    df["_inv_dt"] = pd.to_datetime(df["Invoice Date & Time"], errors="coerce")
    mask = (df["_inv_dt"] >= start_date) & (df["_inv_dt"] <= end_date + " 23:59:59")
    df = df[mask].drop(columns=["_inv_dt"])

    return df


# ---------------------------------------------------------------------------
# Yesterday Completed Trips
# ---------------------------------------------------------------------------

def build_yesterday_completed_trips(mtr: pd.DataFrame, data_date: str, days_back: int = 1) -> pd.DataFrame:
    """days_back=1 (default) reproduces the original, already-validated
    behavior exactly: Stamp Time date == data_date, no range. Set higher
    only for late/catch-up runs (per explicit instruction) -- e.g.
    days_back=3 pulls the 3 days ending at data_date inclusive, covering 2
    additional prior days Khagash's team sometimes has to catch up on when
    a run is late. Confirmed this is a real-world catch-up scenario, not a
    logic bug in the single-day filter -- see module docstring.
    """
    df = mtr.copy()
    df["_stamp_dt"] = pd.to_datetime(df["Stamp Time"], errors="coerce")
    range_start = (pd.to_datetime(data_date) - pd.Timedelta(days=days_back - 1)).strftime("%Y-%m-%d")
    df = df[(df["_stamp_dt"].dt.date.astype(str) >= range_start) & (df["_stamp_dt"].dt.date.astype(str) <= data_date)].copy()

    def nonsto(dc):
        dc = str(dc).strip()
        if dc in ("10", "20"):
            return "NON STO"
        if dc == "30":
            return "STO"
        return None

    df["STO/NON STO"] = df["Distribution Channel"].apply(nonsto)

    def deviation_remark(row):
        # CONFIRMED: uses Destination Deviation, not Lead Distance Deviation
        # (the SOP text names the wrong field — see module docstring).
        try:
            dd = float(row["Destination Deviation"])
        except (TypeError, ValueError):
            return "Trip Reject"
        if dd == 1:
            return "At Destination"
        if 1 < dd <= 20:
            return "Less Than 20 Km"
        if 20 < dd <= 40:
            if row["STO/NON STO"] == "NON STO":
                return "Less Than 40 Km"
            return row["Destination Deviation Direction"]
        return row["Destination Deviation Direction"]

    df["Deviation Remarks"] = df.apply(deviation_remark, axis=1)

    return df.drop(columns=["_stamp_dt"])


# ---------------------------------------------------------------------------
# JKLC Offline Trips
# ---------------------------------------------------------------------------

def build_offline_trips(mtr: pd.DataFrame, offline_dashboard: pd.DataFrame,
                        data_date: str) -> pd.DataFrame:
    # 5-day window ENDING at data_date, per the SOP's own worked example
    # ("today is 10 June -> take 9,8,7,6,5") and confirmed against the real
    # file's window title "08-07-2026 - 12-07-2026" (ends at data_date=12,
    # not the day after) -- see module docstring correction #2.
    last_5_days_start = (pd.to_datetime(data_date) - pd.Timedelta(days=4)).strftime("%Y-%m-%d")
    df = mtr.copy()
    df["_inv_dt"] = pd.to_datetime(df["Invoice Date & Time"], errors="coerce")
    df = df[(df["_inv_dt"] >= last_5_days_start) & (df["_inv_dt"] <= data_date + " 23:59:59")].copy()

    offline_only = offline_dashboard[offline_dashboard["Status"] == "Unreachable"]
    last_seen_map = offline_only.drop_duplicates("Vehicle").set_index("Vehicle")["Last Seen"]

    df["AT Offline"] = df["Vehicle No."].map(last_seen_map)
    df = df[df["AT Offline"].notna()].copy()

    df["_at_offline_dt"] = pd.to_datetime(df["AT Offline"], errors="coerce")

    def classify(row):
        if pd.isna(row["_at_offline_dt"]) or pd.isna(row["_inv_dt"]):
            return None
        if row["_at_offline_dt"] > row["_inv_dt"]:
            return "Vehicle Offline Post Gate Out"
        return "Vehicle Take Load in Offline Condition"

    df["Offline Remarks"] = df.apply(classify, axis=1)

    return df.drop(columns=["_inv_dt", "_at_offline_dt"])


# ---------------------------------------------------------------------------
# All Installation — validated EXACT (8684 = 8684 against real 12th July)
# ---------------------------------------------------------------------------

ALL_INSTALLATION_OUTPUT_COLS = ["Company", "Vehicles", "IMEI", "Mobile Num", "Installation", "Veh Add Date"]


def build_all_installation(device_status: pd.DataFrame, dashboard: pd.DataFrame) -> pd.DataFrame:
    ds = device_status.copy()
    ds["IMEI"] = ds["IMEI"].astype(str)
    dashboard = dashboard.copy()
    dashboard["Imei"] = dashboard["Imei"].astype(str)
    dash_imeis = set(dashboard["Imei"])
    matched = ds[ds["IMEI"].isin(dash_imeis)].copy()

    # CORRECTED: previously kept every raw Device Status column (19 cols:
    # Company Id, Code, Autovid, Group, VehicleId, Port, Recovery Date,
    # First Data Date, Last Data Date, Specifications, Battery,
    # Reflections, Device Type, etc.). Comparing against the real 12th
    # July file showed it only keeps 6 of those, plus a new 1-based row
    # serial ("S. No.") that isn't in the source data at all.
    out = matched[ALL_INSTALLATION_OUTPUT_COLS].reset_index(drop=True)
    out.insert(0, "S. No.", range(1, len(out) + 1))
    return out


def build_month_installation(all_installation: pd.DataFrame, data_date: str) -> pd.DataFrame:
    """SOP step 3: All Installation filtered to the current month.
    Tab is named dynamically by the caller, e.g. "July '26 Installation".
    CONFIRMED: S. No. is independently renumbered 1..N for this subset
    (checked the real file: 74 rows, S. No. goes 1..74) -- not carried over
    from All Installation's own numbering.
    """
    df = all_installation.drop(columns=["S. No."]).copy()
    df["_add_dt"] = pd.to_datetime(df["Veh Add Date"], errors="coerce")
    month_start = pd.to_datetime(data_date).replace(day=1)
    month_end = month_start + pd.offsets.MonthEnd(0)
    mask = (df["_add_dt"] >= month_start) & (df["_add_dt"] <= month_end.strftime("%Y-%m-%d") + " 23:59:59")
    out = df[mask].drop(columns=["_add_dt"]).reset_index(drop=True)
    out.insert(0, "S. No.", range(1, len(out) + 1))
    return out


def month_installation_tab_name(data_date: str) -> str:
    dt = pd.to_datetime(data_date)
    return f"{dt.strftime('%B')} '{dt.strftime('%y')} Installation"


# ---------------------------------------------------------------------------
# Durg Dispatch Status of Yesterday + Summary block
#
# Per the SOP docx (Section 5), NOT just the original script's 4-metric
# version -- 5 more fields added/fixed here, see module docstring.
# ---------------------------------------------------------------------------

def build_durg_dispatch(mtr: pd.DataFrame, device_status: pd.DataFrame,
                        master_entry: pd.DataFrame, all_installation: pd.DataFrame,
                        data_date: str) -> dict:
    df = mtr.copy()
    df["_inv_dt"] = pd.to_datetime(df["Invoice Date & Time"], errors="coerce")
    durg = df[
        (df["Org Location"] == "JKLC Durg")
        & (df["_inv_dt"].dt.date.astype(str) == data_date)
    ].copy()
    # NOT deduped by Vehicle No., despite the SOP text saying to ("Remove
    # duplicate rows by Vehicle No."). Confirmed wrong against the real
    # 12th July file: deduping gave 177 vehicles / 146 AT-matched, but the
    # real report's own numbers are 195 / 163 -- which match this file's
    # RAW (undeduped) row count and AT-match count exactly. A vehicle
    # making 2 dispatches in a day apparently counts twice here.

    at_vehicles = set(device_status["Vehicles"].astype(str))
    durg["AT"] = durg["Vehicle No."].apply(lambda v: v if v in at_vehicles else None)

    wheelseye_vehicles = set(
        master_entry[master_entry["GPS Provider Name"].str.upper() == "WHEELSEYE"]["Vehicle no."].astype(str)
    )
    durg["Wheels Eye"] = durg.apply(
        lambda r: r["Vehicle No."] if (r["AT"] is None and r["Vehicle No."] in wheelseye_vehicles) else None,
        axis=1,
    )

    total_dispatched = len(durg)
    at_matched = int(durg["AT"].notna().sum())
    wheelseye_matched = int(durg["Wheels Eye"].notna().sum())
    no_gps = int(((durg["AT"].isna()) & (durg["Wheels Eye"].isna())).sum())

    # No rule for this anywhere in the SOP docx or script; 0 in every real
    # example seen. Fixed placeholder, not computed — see module docstring.
    clinker = 0

    # SOP 5.6: "go to the All Installation sheet, filter to yesterday's date
    # and Plant = Durg, and count the rows." This is the SAME figure as the
    # top Summary block's "Last Day AT Installation" for Durg (confirmed
    # identical in the real file: both are 6).
    plant_install = all_installation[all_installation["Company"] == "JKLC Durg"].copy()
    plant_install["_add_dt"] = pd.to_datetime(plant_install["Veh Add Date"], errors="coerce")
    allotted_to_axestrack = int((plant_install["_add_dt"].dt.date.astype(str) == data_date).sum())
    axestrack_installation_count = allotted_to_axestrack  # SOP: same count, two names

    # SOP 5.7: NOT derived from Total Dispatched -- it's allotted + no-GPS.
    # Confirmed via real numbers: 6 + 31 = 37, matches exactly.
    total_prospect = allotted_to_axestrack + no_gps

    return {
        "detail": durg,
        "Total Vehicles Dispatched from Plant": total_dispatched,
        "Clinker Vehicle (Out of Prospect for GPS Installations)": clinker,
        "Axestrack Pre-installed for JKLC": at_matched,
        "Wheelseye Pre- Installed for JKLC": wheelseye_matched,
        "Total Prospect Vehicles for GPS installations arrived at Plant": total_prospect,
        "Vehicles allotted to Axestrack": allotted_to_axestrack,
        "Axestrack's Installation Count": axestrack_installation_count,
        "Wheelseye New Installation": 0,  # no rule specified anywhere; 0 in every real example
        "Vehicles allowed by plant without GPS ": no_gps,  # trailing space matches real file's label
    }


# ---------------------------------------------------------------------------
# Summary tab (top block, per-plant)
# ---------------------------------------------------------------------------

def build_summary(yesterday_completed: pd.DataFrame, offline_trips: pd.DataFrame,
                  all_installation: pd.DataFrame, durg_dispatch: dict, data_date: str) -> dict:
    rows = {}
    for plant in ORG_LOCATIONS_ALL:
        g = yesterday_completed[yesterday_completed["Org Location"] == plant]
        backward = int((g["Deviation Remarks"] == "Backward").sum())
        forward = int((g["Deviation Remarks"] == "Forward").sum())

        off = offline_trips[offline_trips["Org Location"] == plant]
        offline_count = len(off)
        post_gate_out = int((off["Offline Remarks"] == "Vehicle Offline Post Gate Out").sum())
        take_load_offline = int((off["Offline Remarks"] == "Vehicle Take Load in Offline Condition").sum())

        plant_install = all_installation[all_installation["Company"] == plant].copy()
        install_total = len(plant_install)
        plant_install["_add_dt"] = pd.to_datetime(plant_install["Veh Add Date"], errors="coerce")
        last_day_install = int((plant_install["_add_dt"].dt.date.astype(str) == data_date).sum())

        # No rule for this anywhere in the SOP docx or script; 0 in every
        # real example seen. Fixed placeholder, not computed.
        third_party = 0

        # CONFIRMED via real file's own numbers (6281+6+0=6287 for Durg).
        jklc_total_install = install_total + last_day_install + third_party

        rows[plant] = {
            "Last Day AT Installation": last_day_install,
            "AT Total Installation Till Date": install_total,
            "3rd Party Vehicle Count": third_party,
            "JKLC Total Installation": jklc_total_install,
            "Backward": backward,
            "Forward": forward,
            "Offline Trips Count": offline_count,
            "Vehicle Offline Post Gate Out": post_gate_out,
            "Vehicle Take Load in Offline Condition": take_load_offline,
        }

    return {"by_plant": rows, "durg_dispatch": durg_dispatch}


# ---------------------------------------------------------------------------
# Excel formatting — extracted cell-by-cell from the real
# JKLC Daily Tracking Report 12th July 2026.xlsx (Summary tab)
# ---------------------------------------------------------------------------

TITLE_FONT = Font(name="Bahnschrift Light", size=9, bold=True)
TITLE_FILL = PatternFill(fill_type="solid", fgColor="F8CBAD")  # theme accent2, lighter 40%

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DAE3F3")  # theme accent1, lighter 80%

LABEL_FONT = Font(bold=True, color="FFFFFF")
LABEL_FILL = PatternFill(fill_type="solid", fgColor="44546A")  # theme dk2

DURG_TITLE_FONT = Font(bold=True)
DURG_TITLE_FILL = PatternFill(fill_type="solid", fgColor="F8CBAD")

DURG_HEADER_FONT = Font(bold=True)
DURG_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")

DURG_LABEL_FONT = Font(bold=True)
DURG_LABEL_FILL = PatternFill(fill_type="solid", fgColor="E7E6E6")

THIN = Side(style="thin")
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

SUMMARY_COL_WIDTHS = {"A": 35.33, "B": 12.33, "C": 10.0, "D": 10.89, "E": 10.55, "F": 11.33, "G": 9.66}


def _write_summary_sheet(wb, summary: dict, data_date: str):
    last_5_days_start = (pd.to_datetime(data_date) - pd.Timedelta(days=4)).strftime("%Y-%m-%d")

    ws = wb.create_sheet("Summary")
    for letter, width in SUMMARY_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    dt = pd.to_datetime(data_date)
    ws.merge_cells("A1:F1")
    ws["A1"] = f"JKLC Daily Tracking Report {dt.strftime('%d-%m-%Y')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = LEFT

    plants_row = ["Plants"] + ORG_LOCATIONS_ALL + ["Total"]
    for col_idx, val in enumerate(plants_row, start=1):
        cell = ws.cell(2, col_idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = LEFT if col_idx == 1 else CENTER

    by_plant = summary["by_plant"]

    def write_metric_row(row_idx, label, key):
        cell = ws.cell(row_idx, 1, value=label)
        cell.font = LABEL_FONT
        cell.fill = LABEL_FILL
        cell.border = THIN_BORDER
        values = [by_plant[p][key] for p in ORG_LOCATIONS_ALL]
        for col_idx, v in enumerate(values, start=2):
            c = ws.cell(row_idx, col_idx, value=v)
            c.border = THIN_BORDER
            c.alignment = CENTER
        total_cell = ws.cell(row_idx, 6, value=sum(values))
        total_cell.border = THIN_BORDER
        total_cell.alignment = CENTER

    write_metric_row(3, "Last Day AT Installation", "Last Day AT Installation")
    ws["G3"] = pd.to_datetime(data_date)
    ws["G3"].number_format = "d-mmm-yy"
    write_metric_row(4, "AT Total Installation Till Date", "AT Total Installation Till Date")
    write_metric_row(5, "3rd Party Vehicle Count", "3rd Party Vehicle Count")
    write_metric_row(6, "JKLC Total Installation", "JKLC Total Installation")

    ws.merge_cells("A7:F7")
    ws["A7"] = "Deviation Count of Completed Trips"
    ws["A7"].font = HEADER_FONT
    ws["A7"].fill = HEADER_FILL

    write_metric_row(8, "Backward", "Backward")
    write_metric_row(9, "Forward", "Forward")

    start_dt = pd.to_datetime(last_5_days_start)
    end_dt = pd.to_datetime(data_date)
    ws.merge_cells("A10:F10")
    ws["A10"] = (f"{start_dt.strftime('%d-%m-%Y')} - {end_dt.strftime('%d-%m-%Y')} "
                 "Dispatch AT Fix Offline Count of pending AI Processing Trips")
    ws["A10"].font = HEADER_FONT
    ws["A10"].fill = HEADER_FILL

    write_metric_row(11, "Offline Trips Count", "Offline Trips Count")
    write_metric_row(12, "Vehicle Offline Post Gate Out", "Vehicle Offline Post Gate Out")
    write_metric_row(13, "Vehicle Take Load in Offline Condition", "Vehicle Take Load in Offline Condition")

    ws.merge_cells("A16:B16")
    ws["A16"] = f"JKLC Durg {end_dt.strftime('%d-%m-%y')}Dispatch Status"
    ws["A16"].font = DURG_TITLE_FONT
    ws["A16"].fill = DURG_TITLE_FILL

    ws["A17"] = "Remarks"
    ws["B17"] = "Count"
    for coord in ("A17", "B17"):
        ws[coord].font = DURG_HEADER_FONT
        ws[coord].fill = DURG_HEADER_FILL

    dd = summary["durg_dispatch"]
    durg_metrics = [
        "Total Vehicles Dispatched from Plant",
        "Clinker Vehicle (Out of Prospect for GPS Installations)",
        "Axestrack Pre-installed for JKLC",
        "Wheelseye Pre- Installed for JKLC",
        "Total Prospect Vehicles for GPS installations arrived at Plant",
        "Vehicles allotted to Axestrack",
        "Axestrack's Installation Count",
        "Wheelseye New Installation",
        "Vehicles allowed by plant without GPS ",
    ]
    for i, metric in enumerate(durg_metrics):
        row_idx = 18 + i
        label_cell = ws.cell(row_idx, 1, value=metric)
        label_cell.font = DURG_LABEL_FONT
        label_cell.fill = DURG_LABEL_FILL
        value_cell = ws.cell(row_idx, 2, value=dd[metric])
        value_cell.font = Font(bold=True)
        value_cell.border = THIN_BORDER

    return ws


MTR_FAMILY_DATE_COLUMNS = {
    "Invoice Date & Time": "m/d/yy h:mm",
    "Org Entry Time": "m/d/yy h:mm",
    "Org Exit Time": "m/d/yy h:mm",
    "Dest Reporting Time": "m/d/yy h:mm",
    "Trip AddTime": "m/d/yy h:mm",
}

# Cosmetic-only, does NOT go through the date_columns machinery (that path
# runs pd.to_datetime() on every entry, which would corrupt a phone-number
# column). Just forces text display so long digit strings like Transporter
# Number never render in scientific notation -- matches what Khagash is
# used to seeing, doesn't affect correctness.
MTR_FAMILY_TEXT_COLUMNS = {"Transporter Number"}
ALL_INSTALLATION_DATE_COLUMNS = {"Veh Add Date": "d-mmm-yy"}

DEFAULT_COL_WIDTH = 13.0

# Per-sheet header look, extracted from the real file. MTR-family sheets
# (MTR / Yesterday Completed Trips) use a pale peach header, unbolded;
# JKLC Offline Trips / All Installation / <Month> Installation / API
# Vehicles use a slightly deeper peach, bold. All Installation and the
# Month tab also use a smaller (10pt) font throughout, per the real file.
SHEET_HEADER_STYLES = {
    "MTR": {"fill": "FBE5D6", "bold": False, "size": 11},
    "Yesterday Completed Trips": {"fill": "FBE5D6", "bold": False, "size": 11},
    "JKLC Offline Trips": {"fill": "FCE4D6", "bold": True, "size": 11},
    "All Installation": {"fill": "FCE4D6", "bold": True, "size": 10},
    "API Vehicles": {"fill": "FCE4D6", "bold": True, "size": 11},
}


def _write_data_sheet(wb, sheet_name: str, df: pd.DataFrame, *, col_widths: dict = None,
                      date_columns: dict = None, header_style_key: str = None, text_columns: set = None):
    """Fast bulk write via ws.append (avoids a slow pandas-then-reload round
    trip), with header styling / column widths / date number-formats matched
    to the real master file. Matched by column NAME (not position), since
    this app's extra computed columns (STO/NON STO, Deviation Remarks, AT
    Offline, Offline Remarks) land in a different order than the real file
    -- data/columns themselves are untouched, this only affects display.

    Per-cell font is only set explicitly for sheets whose real font size
    differs from openpyxl's own default (Calibri 11, which already matches
    every MTR-family/Offline/API sheet) -- looping every data cell across
    MTR's ~6000 rows x 121 cols would be a real perf cost for no visible
    change on those sheets (this pattern already caused a Render 502 once
    on Report 2's large tabs, see that report's history).
    """
    col_widths = col_widths or {}
    date_columns = date_columns or {}
    text_columns = text_columns or set()
    style = SHEET_HEADER_STYLES.get(header_style_key or sheet_name, SHEET_HEADER_STYLES["MTR"])

    ws = wb.create_sheet(sheet_name)
    columns = list(df.columns)

    header_font = Font(name="Calibri", size=style["size"], bold=style["bold"])
    header_fill = PatternFill(fill_type="solid", fgColor=style["fill"])
    needs_data_font = style["size"] != 11
    data_font = Font(name="Calibri", size=style["size"]) if needs_data_font else None

    ws.append(columns)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(name.strip(), DEFAULT_COL_WIDTH)

    date_col_idxs = {i: fmt for i, name in enumerate(columns) if (fmt := date_columns.get(name))}
    text_col_idxs = [i for i, name in enumerate(columns) if name in text_columns]

    # Parse date columns ONCE, vectorized, before the row loop. Calling
    # pd.to_datetime() per-scalar inside the loop (the first version of this
    # function did) has heavy per-call overhead -- on MTR's ~6000 rows x 5
    # date columns that alone measured 400+ seconds, which would time out on
    # Render. Vectorized parsing of the whole column is orders of magnitude
    # faster.
    if date_col_idxs:
        df = df.copy()
        for idx, col_name in enumerate(columns):
            if idx in date_col_idxs:
                df[col_name] = pd.to_datetime(df[col_name], errors="coerce")

    # Track the row index manually rather than querying ws.max_row / ws[...]
    # per row -- both recompute by scanning every cell in the sheet, making
    # that call alone O(n) and the whole loop O(n^2). On MTR's ~6000 rows
    # this measured 200+ seconds; tracking a counter is O(1) per row.
    row_idx = 1
    for row in df.itertuples(index=False):
        row_idx += 1
        ws.append(row)
        if date_col_idxs:
            for idx, fmt in date_col_idxs.items():
                ws.cell(row_idx, idx + 1).number_format = fmt
        for idx in text_col_idxs:
            ws.cell(row_idx, idx + 1).number_format = "@"
        if needs_data_font:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row_idx, col_idx).font = data_font

    return ws


def write_output(mtr, yesterday_completed, offline_trips, all_installation,
                 month_installation, month_tab_name, api_vehicles, summary,
                 data_date, output_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, summary, data_date)
    _write_data_sheet(wb, "MTR", mtr,
                      col_widths=MTR_FAMILY_COL_WIDTHS, date_columns=MTR_FAMILY_DATE_COLUMNS,
                      text_columns=MTR_FAMILY_TEXT_COLUMNS)
    _write_data_sheet(wb, "Yesterday Completed Trips", yesterday_completed,
                      col_widths=MTR_FAMILY_COL_WIDTHS, date_columns=MTR_FAMILY_DATE_COLUMNS,
                      text_columns=MTR_FAMILY_TEXT_COLUMNS)
    _write_data_sheet(wb, "API Vehicles", api_vehicles)
    _write_data_sheet(wb, month_tab_name, month_installation,
                      col_widths=ALL_INSTALLATION_COL_WIDTHS, date_columns=ALL_INSTALLATION_DATE_COLUMNS,
                      header_style_key="All Installation")
    _write_data_sheet(wb, "All Installation", all_installation,
                      col_widths=ALL_INSTALLATION_COL_WIDTHS, date_columns=ALL_INSTALLATION_DATE_COLUMNS)
    _write_data_sheet(wb, "JKLC Offline Trips", offline_trips,
                      col_widths=MTR_FAMILY_COL_WIDTHS, date_columns=MTR_FAMILY_DATE_COLUMNS,
                      text_columns=MTR_FAMILY_TEXT_COLUMNS)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Entry point used by the generic upload -> process -> download route
# ---------------------------------------------------------------------------

def process(input_files: dict, dates: dict, output_dir: Path) -> Path:
    # start_date/end_date range, matching Reports 1 & 2's UI. end_date IS
    # data_date -- the day whose data this report covers (Yesterday
    # Completed Trips, Durg Dispatch, Offline Trips window end, Last Day
    # AT Installation all key off it -- see module docstring's convention
    # correction). start_date only bounds the MTR tab's own date filter.
    start_date = dates["start_date"]
    data_date = dates["end_date"]
    # Optional, defaults to 1 (today's existing behavior, unchanged) -- see
    # build_yesterday_completed_trips's docstring. Only ever set higher for
    # late/catch-up runs; no other tab or report is affected by this value.
    yesterday_days_back = dates.get("yesterday_days_back", 1)
    log.info("Processing JKLC Daily Tracking Report for %s to %s (yesterday_days_back=%d)",
             start_date, data_date, yesterday_days_back)

    mtr_raw = _read_mtr(input_files["mtr_raw"])
    device_status = _load_device_status(input_files["device_status"])
    dashboard = _load_dashboard(input_files["dashboard"])
    offline_dashboard = _load_offline_dashboard(input_files["offline_dashboard"])
    master_entry = _load_master_entry_data(input_files["master_entry"])
    api_vehicles = _load_api_vehicles(input_files["api_vehicles"])

    try:
        mtr = clean_mtr(mtr_raw, start_date, data_date)
        yesterday_completed = build_yesterday_completed_trips(mtr, data_date, yesterday_days_back)
        offline_trips = build_offline_trips(mtr, offline_dashboard, data_date)
        all_installation = build_all_installation(device_status, dashboard)
        month_installation = build_month_installation(all_installation, data_date)
        month_tab_name = month_installation_tab_name(data_date)
        durg_dispatch = build_durg_dispatch(mtr, device_status, master_entry, all_installation, data_date)
        summary = build_summary(yesterday_completed, offline_trips, all_installation, durg_dispatch, data_date)
        api_vehicles = clean_api_vehicles(api_vehicles)
    except KeyError as exc:
        raise ReportProcessingError(
            f"Expected column {exc} not found. Check each file was uploaded to the "
            "correct slot (MTR Raw / Device Status / Dashboard / Offline Trip Dashboard / "
            "Master Entry Data / API Vehicles)."
        ) from exc

    log.info(
        "Yesterday Completed: %d | Offline Trips: %d | All Installation: %d | %s: %d | Durg Dispatch: %s",
        len(yesterday_completed), len(offline_trips), len(all_installation),
        month_tab_name, len(month_installation), durg_dispatch,
    )

    output_path = output_dir / f"JKLC_Daily_Tracking_Report_{data_date}.xlsx"
    write_output(mtr, yesterday_completed, offline_trips, all_installation,
                 month_installation, month_tab_name, api_vehicles, summary,
                 data_date, output_path)
    return output_path


register(
    ReportMeta(
        id="3",
        name="JKLC Daily Tracking Report",
        input_slots=[
            InputSlot(key="mtr_raw", label="MTR Raw", accept=".csv,.xlsx",
                     hint="mtr - <timestamp>.csv"),
            InputSlot(key="device_status", label="Device Status (AT portal)", accept=".csv",
                     hint="Device_Status (67).csv"),
            InputSlot(key="dashboard", label="Dashboard Export (AT portal)", accept=".xlsx",
                     hint="dashboard_export_<id>.xlsx"),
            InputSlot(key="offline_dashboard", label="Offline Trip Dashboard (AT portal)", accept=".xls,.html",
                     hint="offline trip Dashboard_p2071_<id>.xls"),
            InputSlot(key="master_entry", label="Master Entry Data (Durg Plant)", accept=".xlsx",
                     hint="Master Entry Data - Durg Plant <date> - Format 2026,wheelseye.xlsx"),
            InputSlot(key="api_vehicles", label="API Vehicles", accept=".csv",
                     hint="api vehicles-<id>.csv"),
        ],
        output_pattern="JKLC_Daily_Tracking_Report_<date>.xlsx (7 tabs: Summary, MTR, Yesterday Completed "
                      "Trips, JKLC Offline Trips, All Installation, <Month> Installation, API Vehicles)",
        process_fn=process,
        implemented=True,
        date_mode="range",
        extra_number_fields=[
            ExtraNumberField(
                key="yesterday_days_back",
                label="Days back (Yesterday Completed Trips only)",
                default=1,
                min_value=1,
                hint=(
                    "1 = yesterday only (normal case, default). Higher = pulls that many days "
                    "ending at End Date -- e.g. 3 covers the 3 days up to and including End Date, "
                    "for late/catch-up runs. Only affects the Yesterday Completed Trips tab; every "
                    "other tab is unaffected."
                ),
            ),
        ],
        notes=(
            "End Date = the day this report covers (matches real filename exactly, e.g. enter "
            "2026-07-12 for the '12th July' report). Start Date only bounds the MTR tab's own date "
            "filter (previously MTR was passed through unfiltered). No cross-day accumulation. "
            "Validated against the real 12th July file: All Installation, Offline Trips (251/206/45), "
            "every AT-installation Summary metric, and Durg Dispatch (195 total, 163 AT-matched, "
            "off by 1 on Wheelseye/No-GPS) all match almost exactly. Durg Dispatch is NOT deduped by "
            "Vehicle No. despite the SOP text saying to -- confirmed the real report double-counts a "
            "vehicle dispatched twice in a day. Deviation Backward/Forward differs from the real file "
            "because it's a multi-day cumulative master (documented, not a bug)."
        ),
    )
)
