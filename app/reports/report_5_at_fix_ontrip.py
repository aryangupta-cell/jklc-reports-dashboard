"""
JKLC AT Fix On-Trip Vehicle Status — web app adaptation (Report 5 of 9)
=========================================================================

Adapted from the provided `jklc_at_fix_ontrip_status.py` script, verified
against the real 12th July 2026 output and its companion reference doc.

This report is a downstream filter of Report 3's own output, not a fresh
build from source data: its only input is the "JKLC Daily Tracking Report"
file, specifically the "JKLC Offline Trips" tab within it.

Validated (confirmed against real data, not guesses):
  1. Filter rule -- CONFIRMED EXACT (38/38 rows against the real 12th July
     output). Filter is Invoice Date & Time == REPORT_DATE, the label date
     ITSELF (e.g. "12th July" report -> filter to 12 July) -- NOT
     REPORT_DATE - 1 despite what the naming might suggest. Independently
     re-verified: filtering Report 3's real Offline Trips tab (252 rows) to
     Invoice Date == 2026-07-12 gives exactly 38 rows; == 2026-07-11 gives
     60. 38 matches the real file's row count exactly.
  2. Column layout -- same columns as the input Offline Trips tab, just
     reordered: "AT Offline", "Offline Remarks", and "Technician Remarks"
     move to sit right after "Org Location" instead of staying at the end.
  3. Sheet2 pivot structure -- one row per plant (aggregate), one row per
     transporter within that plant, then a Grand Total row. Confirmed
     against the real file down to individual transporter subtotals.

Formatting extracted cell-by-cell from the real file (same approach as
Reports 2 & 3): Sheet1 header fill/bold, Sheet2's header/Grand Total rows
(dark blue fill, white text) and plant-aggregate rows (orange fill),
transporter rows unstyled.

TECHNICIAN REMARKS -- NOT YET INDEPENDENTLY VALIDATED, flagging this
clearly rather than presenting it as confirmed:
  Originally this column was left blank on purpose -- confirmed manual
  (WhatsApp replies typed in after generation, 25/38 blank in the real
  12th July file, one "Odd Hour" case fell inside the stated shift window,
  proving it wasn't a computed rule). Khagash's team now maintains a
  Google Sheet ("GPS Remarks" tab: GPS Offline Vehicle No, Remarks, Date)
  that a technician fills in going forward, so this is now joinable.

  Join key is Vehicle No. + Date TOGETHER, not Vehicle No. alone -- per
  explicit instruction, since the same vehicle can go offline on different
  days with different remarks, and matching on Vehicle No. only risks
  pulling the wrong day's remark. The report-side date used for this join
  is REPORT_DATE itself (every row in Sheet1 already shares that date,
  since Sheet1 is pre-filtered to Invoice Date & Time == REPORT_DATE --
  see point 1 above -- so there's no separate per-row event date to use
  instead). This is an ASSUMPTION, not yet confirmed: the sheet only had 2
  sample rows (mostly blank template) at the time this was built, not
  enough to validate the join against real data. In particular, it's not
  yet confirmed whether the sheet's Date column is the actual offline
  EVENT date vs. the date the technician happened to type the remark in --
  if it's the latter, joining on REPORT_DATE could silently miss real
  matches. Re-verify a handful of real Vehicle No. + Date pairs once
  the sheet has genuine filled-in data before trusting this in production.

I/O differences from the original script:
  - Input: 2 uploaded files -- Report 3's output .xlsx, plus the GPS
    Remarks sheet export -- instead of a local path.
  - Output: 1 .xlsx with Sheet1 + Sheet2, streamed back in the same request.
  - Date is a single "Report Date" field (this report is inherently a
    single-day filter, same reasoning as Report 3's date field).
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reports.errors import ReportProcessingError, describe_column_mismatch
from reports.registry import InputSlot, ReportMeta, register

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

# Columns that move to the front, right after Org Location (confirmed order
# from the real output). The Offline Remarks column name is resolved
# separately (see _pick_offline_remarks_col) to handle both the real file's
# " Offline Remarks" (leading space, a quirk of that spreadsheet) and our
# own Report 3 output's "Offline Remarks" (no leading space).
FRONT_COLUMNS_BASE = ["Trip ID", "Invoice no", "Vehicle No.", "Transporter",
                     "Invoice Date & Time", "Org Location", "AT Offline"]

REMARK_CATEGORIES = ["Vehicle Offline Post Gate Out", "Vehicle Take Load in Offline Condition"]


# ---------------------------------------------------------------------------
# File I/O
# ---------------------------------------------------------------------------

def _load_offline_trips(path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name="JKLC Offline Trips", dtype=str)
    except ValueError as exc:
        raise ReportProcessingError(
            f"Couldn't find a 'JKLC Offline Trips' tab in '{path.name}'. "
            "Check you uploaded Report 3's own output file (JKLC Daily Tracking Report)."
        ) from exc
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read '{path.name}': {exc}") from exc


GPS_REMARKS_COLS = ["GPS Offline Vehicle No", "Remarks", "Date"]


def _load_gps_remarks(path: Path) -> pd.DataFrame:
    """Load the technician's GPS Remarks sheet export (Google Sheet ->
    xlsx/csv). Accepts either format since a Google Sheet export could
    arrive as either."""
    try:
        if path.suffix.lower() == ".csv":
            df = pd.read_csv(path, dtype=str)
        else:
            try:
                df = pd.read_excel(path, sheet_name="GPS Remarks", dtype=str)
            except ValueError:
                # Fall back to the first sheet if it isn't literally named
                # "GPS Remarks" (e.g. a re-saved/renamed export).
                df = pd.read_excel(path, dtype=str)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read GPS Remarks file '{path.name}': {exc}") from exc

    mismatch = describe_column_mismatch(df.columns, GPS_REMARKS_COLS, path.name)
    if mismatch:
        raise ReportProcessingError(
            f"{mismatch} Check you uploaded the GPS Remarks sheet export (tab: GPS Remarks)."
        )
    return df[GPS_REMARKS_COLS].copy()


def build_remarks_lookup(gps_remarks: pd.DataFrame) -> dict:
    """Builds a {(vehicle_no, date_str): remark} lookup. Join key is
    Vehicle No. + Date TOGETHER (not vehicle alone) -- the same vehicle can
    go offline on different days with different remarks, see module
    docstring for the full reasoning and the open validation question on
    what the sheet's Date column actually represents.
    """
    df = gps_remarks.copy()
    df["_vehicle"] = df["GPS Offline Vehicle No"].astype(str).str.strip().str.upper()
    df["_date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date.astype(str)
    df = df[df["_vehicle"].notna() & (df["_vehicle"] != "") & (df["_date"] != "NaT")]
    df = df[df["Remarks"].notna() & (df["Remarks"].astype(str).str.strip() != "")]
    # Last one wins if there's ever a genuine duplicate (vehicle, date) pair.
    return {(v, d): r for v, d, r in zip(df["_vehicle"], df["_date"], df["Remarks"])}


def _pick_offline_remarks_col(columns) -> str:
    if " Offline Remarks" in columns:
        return " Offline Remarks"
    if "Offline Remarks" in columns:
        return "Offline Remarks"
    raise ReportProcessingError(
        "Couldn't find an 'Offline Remarks' column in the uploaded file's "
        "JKLC Offline Trips tab."
    )


# ---------------------------------------------------------------------------
# Sheet1: filter + reorder
# ---------------------------------------------------------------------------

def build_sheet1(offline_trips: pd.DataFrame, report_date: str, remarks_lookup: dict) -> pd.DataFrame:
    df = offline_trips.copy()
    remark_col = _pick_offline_remarks_col(df.columns)

    df["_inv_dt"] = pd.to_datetime(df["Invoice Date & Time"], errors="coerce")
    df = df[df["_inv_dt"].dt.date.astype(str) == report_date].copy()
    df = df.drop(columns=["_inv_dt"])

    # Joined from the GPS Remarks sheet on Vehicle No. + Date TOGETHER (not
    # vehicle alone) -- see module docstring for why, and for the open
    # question on what the sheet's Date column actually represents (not yet
    # validated against real filled-in data). Every row here already shares
    # report_date (see the filter above), so that's the date used for every
    # row's half of the join key. Rows with no match stay blank -- not
    # every offline vehicle will have a technician remark yet.
    vehicle_key = df["Vehicle No."].astype(str).str.strip().str.upper()
    df["Technician Remarks"] = [remarks_lookup.get((v, report_date)) for v in vehicle_key]

    front_columns = FRONT_COLUMNS_BASE + [remark_col]
    remaining_cols = [c for c in df.columns if c not in front_columns and c != "Technician Remarks"]
    ordered_cols = front_columns + ["Technician Remarks"] + remaining_cols
    ordered_cols = [c for c in ordered_cols if c in df.columns]

    return df[ordered_cols]


# ---------------------------------------------------------------------------
# Sheet2: pivot
# ---------------------------------------------------------------------------

def build_sheet2_pivot(sheet1: pd.DataFrame):
    """Returns (pivot_df, row_kinds) -- row_kinds is a parallel list of
    "plant" / "transporter" / "total" tags used by the Excel writer to pick
    the right styling per row (plant rows get an orange fill, the Grand
    Total row gets the same dark-blue fill as the header, transporter rows
    are unstyled)."""
    df = sheet1.copy()
    remark_col = _pick_offline_remarks_col(df.columns)

    rows = []
    row_kinds = []
    grand_total = {c: 0 for c in REMARK_CATEGORIES}

    for plant, plant_group in df.groupby("Org Location"):
        plant_counts = plant_group[remark_col].value_counts()
        plant_row = {"Plant Name": plant}
        plant_total = 0
        for c in REMARK_CATEGORIES:
            v = int(plant_counts.get(c, 0))
            plant_row[c] = v if v else None
            plant_total += v
            grand_total[c] += v
        plant_row["Grand Total"] = plant_total
        rows.append(plant_row)
        row_kinds.append("plant")

        for transporter, t_group in plant_group.groupby("Transporter"):
            t_counts = t_group[remark_col].value_counts()
            t_row = {"Plant Name": transporter}
            t_total = 0
            for c in REMARK_CATEGORIES:
                v = int(t_counts.get(c, 0))
                t_row[c] = v if v else None
                t_total += v
            t_row["Grand Total"] = t_total
            rows.append(t_row)
            row_kinds.append("transporter")

    total_row = {"Plant Name": "Grand Total"}
    grand_sum = 0
    for c in REMARK_CATEGORIES:
        total_row[c] = grand_total[c]
        grand_sum += grand_total[c]
    total_row["Grand Total"] = grand_sum
    rows.append(total_row)
    row_kinds.append("total")

    pivot_df = pd.DataFrame(rows, columns=["Plant Name"] + REMARK_CATEGORIES + ["Grand Total"])
    return pivot_df, row_kinds


# ---------------------------------------------------------------------------
# Formatting -- extracted cell-by-cell from the real
# JKLC AT Fix Ontrip Vehicles Status- 12th July 2026.xlsx
# ---------------------------------------------------------------------------

SHEET1_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
SHEET1_HEADER_FONT = Font(name="Calibri", size=11, bold=True)
SHEET1_DATA_FONT = Font(name="Calibri", size=11)

# theme accent1 (4472C4) tint -0.5 -> dark blue, for header/Grand Total rows.
PIVOT_HEADER_FILL = PatternFill(fill_type="solid", fgColor="203864")
PIVOT_HEADER_FONT = Font(name="Calibri", size=11, color="FFFFFF")
# theme accent2 (ED7D31) tint 0.4 -> medium orange, for plant-aggregate rows.
PIVOT_PLANT_FILL = PatternFill(fill_type="solid", fgColor="F4B183")
PIVOT_DATA_FONT = Font(name="Calibri", size=11)

THIN = Side(style="thin")
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

SHEET1_COL_WIDTHS = {
    "Trip ID": 10.89, "Invoice no": 14.22, "Vehicle No.": 15.0, "Transporter": 42.55,
    "Invoice Date & Time": 22.55, "Org Location": 16.11, "AT Offline": 15.44,
    "Offline Remarks": 32.44, " Offline Remarks": 32.44, "Technician Remarks": 57.66,
}
DEFAULT_SHEET1_COL_WIDTH = 15.0

PIVOT_COL_WIDTHS = {"Plant Name": 43.22, "Vehicle Offline Post Gate Out": 25.78,
                    "Vehicle Take Load in Offline Condition": 33.33, "Grand Total": 10.55}


def _write_sheet1(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Sheet1")
    columns = list(df.columns)

    ws.append(columns)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx)
        cell.font = SHEET1_HEADER_FONT
        cell.fill = SHEET1_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = SHEET1_COL_WIDTHS.get(name.strip(), DEFAULT_SHEET1_COL_WIDTH)

    for row in df.itertuples(index=False):
        ws.append(row)

    return ws


def _write_sheet2(wb, df: pd.DataFrame, row_kinds: list):
    """Written starting at row 3 (2 blank rows above), matching the real
    file and the original script's own startrow=2 convention."""
    ws = wb.create_sheet("Sheet2")
    columns = list(df.columns)
    header_row = 3

    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(header_row, col_idx, value=name)
        cell.font = PIVOT_HEADER_FONT
        cell.fill = PIVOT_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = PIVOT_COL_WIDTHS.get(name, 15.0)

    for row_offset, (row, kind) in enumerate(zip(df.itertuples(index=False), row_kinds), start=1):
        row_idx = header_row + row_offset
        if kind == "total":
            font, fill = PIVOT_HEADER_FONT, PIVOT_HEADER_FILL
        elif kind == "plant":
            font, fill = PIVOT_DATA_FONT, PIVOT_PLANT_FILL
        else:
            font, fill = PIVOT_DATA_FONT, None

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value=value)
            cell.font = font
            if fill is not None:
                cell.fill = fill
            cell.alignment = CENTER

    return ws


# ---------------------------------------------------------------------------
# Entry point used by the generic upload -> process -> download route
# ---------------------------------------------------------------------------

def process(input_files: dict, dates: dict, output_dir: Path) -> Path:
    report_date = dates["report_date"]
    log.info("Processing JKLC AT Fix On-Trip Vehicle Status for %s", report_date)

    offline_trips = _load_offline_trips(input_files["daily_tracking_output"])
    log.info("Input JKLC Offline Trips rows: %d", len(offline_trips))

    gps_remarks = _load_gps_remarks(input_files["gps_remarks_sheet"])
    remarks_lookup = build_remarks_lookup(gps_remarks)
    log.info("GPS Remarks rows: %d | usable (vehicle+date+remark all present): %d",
             len(gps_remarks), len(remarks_lookup))

    try:
        sheet1 = build_sheet1(offline_trips, report_date, remarks_lookup)
    except KeyError as exc:
        raise ReportProcessingError(
            f"Expected column {exc} not found. Check you uploaded Report 3's own output file "
            "(JKLC Daily Tracking Report) with its JKLC Offline Trips tab intact."
        ) from exc

    log.info("Sheet1 (filtered to %s): %d rows", report_date, len(sheet1))
    sheet2, row_kinds = build_sheet2_pivot(sheet1)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet2(wb, sheet2, row_kinds)
    _write_sheet1(wb, sheet1)

    output_path = output_dir / f"JKLC_AT_Fix_Ontrip_Vehicles_Status_{report_date}.xlsx"
    wb.save(output_path)
    return output_path


def process_dispatch(input_files: dict, dates: dict, output_dir: Path) -> Path:
    """Entry point wired into the registry below. Offloads to the office
    server over SSH when SSH_HOST is configured (see core/ssh_worker.py +
    office_server_worker.py at the repo root), so this report also runs on
    the office server's CPU/RAM instead of Render's. With no office server
    configured (the default), falls straight through to the same
    `process()` above unchanged -- no change to this report's actual logic
    or output, just where it executes.
    """
    from core.ssh_worker import is_configured, run_remote

    if is_configured():
        return run_remote("5", input_files, dates, output_dir)
    return process(input_files, dates, output_dir)


register(
    ReportMeta(
        id="5",
        name="JKLC AT Fix On-Trip Vehicle Status",
        input_slots=[
            InputSlot(
                key="daily_tracking_output",
                label="Daily Tracking Report Output",
                accept=".xlsx",
                hint="JKLC_Daily_Tracking_Report_<date>.xlsx",
            ),
            InputSlot(
                key="gps_remarks_sheet",
                label="GPS Remarks Sheet (technician export)",
                accept=".xlsx,.csv",
                hint="GPS Remarks tab: GPS Offline Vehicle No, Remarks, Date",
            ),
        ],
        output_pattern="JKLC_AT_Fix_Ontrip_Vehicles_Status_<date>.xlsx (Sheet1 filtered/reordered data, "
                      "Sheet2 plant/transporter pivot)",
        process_fn=process_dispatch,
        implemented=True,
        date_mode="single",
        notes=(
            "Downstream filter of Report 3's own output -- only input is Report 3's 'JKLC Offline "
            "Trips' tab. Report Date filter is the label date ITSELF (e.g. enter 2026-07-12 for the "
            "'12th July' report), NOT date-1, despite the naming -- confirmed exact (38/38 rows) "
            "against the real file. Technician Remarks is now joined from the GPS Remarks sheet on "
            "Vehicle No. + Report Date together (NOT vehicle alone, since a vehicle can go offline on "
            "different days with different remarks) -- rows with no match stay blank. NOT YET "
            "INDEPENDENTLY VALIDATED against real filled-in data (the sheet had only 2 sample rows "
            "when this was built) -- in particular, unconfirmed whether the sheet's Date column is "
            "the actual offline event date or just when the technician typed the remark; re-verify "
            "real Vehicle No. + Date pairs once genuine data exists before trusting this live."
        ),
    )
)
