"""
Control Tower Tracker (Backward Deviation) — web app adaptation (Report 7 of 9)
==================================================================================

Adapted from the provided `jklc_control_tower_tracker.py` script and its
companion reference doc, independently re-verified before wiring in.

Rebuilds the cumulative "Control_Tower_Tracker_<date>.xlsx" workbook (2 tabs:
Base, Summary) Khagash maintains by hand. Same growing-history pattern as
Report 6's Master workbook: takes the PREVIOUS Control Tower Tracker file,
adds a new day's "Backward" candidates from Report 3's Yesterday Completed
Trips tab, applies confirmed automatic removal rules, and appends survivors.

Confirmed business logic (validated against two real Control Tower Tracker
snapshots -- 2026-07-13 and 2026-07-01, 391 real rows total; see reference
doc for the original evidence trail):

  1. Candidates = Yesterday Completed Trips rows where Deviation Remarks ==
     "Backward".
  2. Rule 8 (dedup by INVNO) -- confirmed 0 duplicates across 351 real rows.
     A candidate whose INVNO already exists anywhere in Base is dropped.
  3. Automatic removal rules apply ONLY to candidates still tagged
     "Backward" -- CRITICAL REFINEMENT found during validation: never touch
     a row already manually reclassified to "Hit" or "Data Missed". Human
     classification always overrides these rules.
       a. Rule 2 -- Transportation Zone == Trans Zone Actual -> remove.
          Confirmed: the one real candidate meeting this was removed.
       b. Rule 3 -- Track Health % < 50 -> remove. Implemented per literal
          spec; NOT YET CONFIRMED with a positive real trigger case.
       c. Rule 4 -- 'RMC' in SHIP TO NM (case-insensitive) -> remove. NOT
          YET CONFIRMED -- 0 real examples in either validation file.
       d. Rule 6 -- dump code (SOLD TO, leading D stripped) matches the
          "Package & SIM Depot" list -> remove. Confirmed: 9/14 real
          removed rows matched exactly.
       e. Rule 7 -- dump code matches "Deviation Upto 40 Km" list AND
          Destination Deviation < 41 -> remove. If code matches but
          deviation >= 41, left for manual review, NOT auto-removed.
          Confirmed: 4/14 real removed rows matched exactly.
       f. Rule 5 (SOLD TO D-stripping) -- CONFIRMED UNIVERSAL, stronger
          than the brief's literal wording ("DG and DD" only). Real data
          shows ANY D<letter> prefix gets stripped (DK11->K11, DJ04->J04,
          DO01->O01, etc.) -- applied to the OUTPUT, not just internal
          rule matching.
       g. STO/NON STO -- already computed upstream in MTR/Yesterday
          Completed Trips; carried through as-is, not recomputed.
  4. Date/check Date columns: Date = the day being reviewed (source data's
     day), check Date = Date + 1 (the day the review ran). The only 2
     columns in the 111-column Base tab not already in Yesterday Completed
     Trips.
  5. Summary tab: carried forward, PLUS a new blank date-block appended
     for this run's report_date (see _add_summary_date_block) -- confirmed
     Summary is a manual data-entry table (Khagash types in her own daily
     counts per date-block by hand), not computed from Base, so this only
     scaffolds a new block ready for her to fill in; it doesn't compute
     the actual counts. Idempotent -- re-running for a date that already
     has a block does nothing extra.

INDEPENDENTLY RE-VERIFIED (not just taking the provided script/doc's word
for it): reconstructed a synthetic "previous" tracker by removing the real
July 12 group (26 rows) from Control_Tower_Tracker_13_July_26.xlsx, then
re-ran with the real July 12 Daily Tracking Report + real dump code lists.
Result: 80/80 rows, identical INVNO sets, removal breakdown 9+4+1=14 exact
match. Diffed all 111 columns cell-by-cell: only genuine difference was 7
rows' Deviation Remarks (the expected manual Backward->Hit reclassifications
this script correctly does not attempt to replicate) -- 0 other diffs, after
correcting for a false-positive from naive NaN-string comparison in my own
diff script (not a real data issue).

NOT automated -- deliberately, not guessed (see reference doc §4):
  - Rule 1 (HIT classification): confirmed manual (Khagash's own notes:
    Google Maps check, "2hrs for 40 cases"). No computable pattern found.
  - Rule 9 (breakdown removal via mail): no mail data source in this
    pipeline.
  - Open edge case, not resolved here: if a human manually deletes a row
    from Base to reflect a confirmed breakdown, and the same invoice later
    resurfaces as a fresh "Backward" candidate, rule 8's dedup would NOT
    re-exclude it (it checks "does this INVNO exist in the previous Base",
    which is false once deleted). Unlikely scenario, flagged not silently
    assumed handled.

Formatting matched cell-by-cell from the real Control Tower Tracker
13_July_26.xlsx: header fill (theme accent2 tint 0.8), Aptos Narrow 12pt
font (both header and data rows -- unlike some other reports here, this
real file uses the SAME font for both), thin borders, centered alignment,
and per-column widths.

I/O differences from the original script:
  - Inputs: 4 uploaded files (previous tracker .xlsx, Report 3's own output
    .xlsx, Package & SIM Depot list, Deviation Upto 40 Km list) instead of
    CLI paths.
  - Output: streamed back in the same request instead of a local --output
    path.
  - Report Date is a single web-form date field (this report is inherently
    a single-day increment, same reasoning as Reports 3, 5 & 6).
"""

import logging
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reports._report7_col_widths import BASE_COL_WIDTHS
from reports.errors import ReportProcessingError
from reports.registry import InputSlot, ReportMeta, register

log = logging.getLogger(__name__)

NEW_COLS = ["Date", "check Date"]


# ---------------------------------------------------------------------------
# File I/O
# ---------------------------------------------------------------------------

def _load_dump_code_set(path: Path, code_col: str = "Dump Code") -> set:
    try:
        df = pd.read_excel(path, sheet_name="Sheet1")
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read '{path.name}': {exc}") from exc
    if code_col not in df.columns:
        raise ReportProcessingError(f"'{path.name}' is missing the expected '{code_col}' column.")
    return set(df[code_col].dropna().astype(str).str.strip())


def _load_prev_base(path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name="Base")
    except Exception as exc:
        raise ReportProcessingError(
            f"Couldn't read '{path.name}': {exc}. Check you uploaded the previous day's "
            "Control_Tower_Tracker_<date>.xlsx (needs a Base tab)."
        ) from exc


def _load_yesterday_completed_trips(path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name="Yesterday Completed Trips")
    except Exception as exc:
        raise ReportProcessingError(
            f"Couldn't read '{path.name}': {exc}. Check you uploaded Report 3's own output file "
            "(JKLC Daily Tracking Report) with its Yesterday Completed Trips tab intact."
        ) from exc


# ---------------------------------------------------------------------------
# Removal rules
# ---------------------------------------------------------------------------

_DPREFIX_PATTERN = re.compile(r"^D[A-Z]")


def apply_removal_rules(candidates: pd.DataFrame, package_sim_codes: set, deviation_40km_codes: set):
    """Apply rules 2, 3, 4, 6, 7 to 'Backward'-tagged candidates only.
    Also applies the confirmed SOLD TO 'strip leading D' transform to
    survivors (rule 5) -- CONFIRMED UNIVERSAL against real data: every
    D+letter-prefixed SOLD TO code (not just 'DG'/'DD' as the brief's
    wording implied) has its leading D removed in the real output.
    Returns (survivors_df, removal_reason_counts dict) for transparency.
    """
    df = candidates.copy()
    reasons = pd.Series([None] * len(df), index=df.index)

    zone_match = df["Transportation Zone"] == df["Trans Zone Actual"]
    reasons[zone_match & reasons.isna()] = "rule2_zone_match"

    low_health = pd.to_numeric(df["Track Health %"], errors="coerce") < 50
    reasons[low_health & reasons.isna()] = "rule3_low_track_health"

    is_rmc = df["SHIP TO NM"].astype(str).str.contains("rmc", case=False, na=False)
    reasons[is_rmc & reasons.isna()] = "rule4_rmc_site"

    stripped_code = df["SOLD TO"].astype(str).str.strip().apply(
        lambda s: s[1:] if s.startswith("D") and len(s) > 1 else s
    )

    in_package_sim = stripped_code.isin(package_sim_codes)
    reasons[in_package_sim & reasons.isna()] = "rule6_package_sim_match"

    in_40km = stripped_code.isin(deviation_40km_codes)
    under_41 = pd.to_numeric(df["Destination Deviation"], errors="coerce") < 41
    reasons[in_40km & under_41 & reasons.isna()] = "rule7_40km_allowance_match"
    # in_40km & NOT under_41 -> "investigate further", NOT auto-removed

    df["_removal_reason"] = reasons
    survivors = df[df["_removal_reason"].isna()].drop(columns=["_removal_reason"]).copy()

    survivors["SOLD TO"] = survivors["SOLD TO"].astype(str).apply(
        lambda s: s[1:] if _DPREFIX_PATTERN.match(s) else s
    )

    removed_counts = df.loc[df["_removal_reason"].notna(), "_removal_reason"].value_counts().to_dict()
    return survivors, removed_counts


def build_new_base_rows(prev_base: pd.DataFrame, yct: pd.DataFrame, package_sim_codes: set,
                        deviation_40km_codes: set, report_date: pd.Timestamp):
    candidates = yct[yct["Deviation Remarks"] == "Backward"].copy()
    raw_count = len(candidates)

    candidates["INVNO"] = pd.to_numeric(candidates["INVNO"], errors="coerce").astype("Int64")
    already_checked = set(prev_base["INVNO"].dropna().astype("int64"))
    before_dedup = len(candidates)
    candidates = candidates[~candidates["INVNO"].isin(already_checked)]
    dropped_dedup = before_dedup - len(candidates)

    survivors, removed_counts = apply_removal_rules(candidates, package_sim_codes, deviation_40km_codes)

    check_date = report_date + pd.Timedelta(days=1)
    common_cols = [c for c in prev_base.columns if c in survivors.columns or c in NEW_COLS]
    new_rows = pd.DataFrame(index=survivors.index)
    for c in common_cols:
        if c == "Date":
            new_rows[c] = report_date
        elif c == "check Date":
            new_rows[c] = check_date
        elif c in survivors.columns:
            new_rows[c] = survivors[c]
        else:
            new_rows[c] = None
    new_rows = new_rows[prev_base.columns]

    stats = {
        "raw_candidates": raw_count,
        "dropped_already_checked": dropped_dedup,
        "removed_by_rule": removed_counts,
        "new_rows_added": len(new_rows),
    }
    return new_rows, stats


# ---------------------------------------------------------------------------
# Formatting -- extracted cell-by-cell from the real
# Control Tower Tracker 13_July_26.xlsx
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FBE3D6")  # theme accent2 tint 0.8
BASE_FONT_NOT_BOLD = Font(name="Aptos Narrow", size=12, bold=False)
THIN = Side(style="thin")
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
DEFAULT_COL_WIDTH = 13.0


# Cosmetic-only: forces text display so long digit strings like Transporter
# Number never render in scientific notation -- matches what Khagash is used
# to seeing, doesn't affect correctness.
TEXT_FORMAT_COLUMNS = {"Transporter Number"}


_SUMMARY_SPOC_FONT = Font(name="Arial", size=12, bold=True)
_SUMMARY_SPOC_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_SUMMARY_DATE_FILL = PatternFill(fill_type="solid", fgColor="FFE699")
_SUMMARY_TOP_CENTER = Alignment(horizontal="center", vertical="top")
_SUMMARY_THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _add_summary_date_block(wb, report_date: pd.Timestamp):
    """Appends a new 5-column date-block to the Summary tab for
    `report_date` -- confirmed cell-by-cell against real data spanning ~30
    existing blocks that Summary is a MANUAL data-entry table (Khagash
    types in her own daily counts by hand), not something computed from
    Base -- so this only scaffolds a new block ready for her to fill in,
    matching the real file's exact repeating structure:
      - Row 1 ("Spoc"): "Khagash", Arial 12 bold, light-blue (D9E1F2) fill,
        thin border, center/top aligned -- first column of the block only.
      - Row 2 ("Date Of Reporting"): report_date, same font/border/align,
        yellow (FFE699) fill, "d-mmm-yy" number format -- first column only.
      - Row 3 ("Plant Name"): Surat / Cuttack / Durg / Jharli / Grand Total
        headers across all 5 columns, same font/align, no fill/border.
      - Rows 4-14 (the metric rows): Surat/Cuttack/Durg/Jharli cells left
        BLANK for manual entry; Grand Total column gets a formula summing
        the other 4 (confirmed exact reference order from real cells, e.g.
        "=E4+D4+C4+B4" -- Jharli+Durg+Cuttack+Surat, right-to-left); same
        font/align on every cell, no fill/border.
      - Row 15 ("% of Deviation"): literal 1 in all 5 columns, "0%" number
        format (matches real file -- every existing cell here is a
        hardcoded 1, not a formula, in every block checked).

    Idempotent: if a block for this exact date already exists anywhere in
    Summary, does nothing -- re-running for the same date (e.g. a
    same-day correction re-run) doesn't duplicate columns.
    """
    if "Summary" not in wb.sheetnames:
        return
    ws = wb["Summary"]
    max_col = ws.max_column

    for start_col in range(2, max_col + 1, 5):
        existing = ws.cell(2, start_col).value
        if existing is not None:
            try:
                if pd.Timestamp(existing).normalize() == report_date.normalize():
                    return
            except (ValueError, TypeError):
                pass

    new_start = max_col + 1 if max_col >= 2 else 2
    surat_c, cuttack_c, durg_c, jharli_c, total_c = range(new_start, new_start + 5)
    block_cols = (surat_c, cuttack_c, durg_c, jharli_c, total_c)

    last_col_letter = get_column_letter(new_start + 4)
    first_col_letter = get_column_letter(new_start)

    spoc_cell = ws.cell(1, surat_c, "Khagash")
    date_cell = ws.cell(2, surat_c, report_date.to_pydatetime())

    # Every real block merges the Spoc/Date Of Reporting cells across all 5
    # columns (e.g. "B1:F1") -- this was missing entirely before, which is
    # why the date showed as "######" (crammed into one narrow column
    # instead of the merged 5-column width) and the border only ever
    # touched the single first cell instead of the whole visual block.
    #
    # Merge FIRST, then style every cell in the range -- verified this
    # exact order in isolation: merging first and then assigning border to
    # each cell (including the now-MergedCell ones) correctly persists a
    # proper box border through save/reload, with each cell keeping only
    # the side(s) relevant to its position (e.g. a middle cell keeps just
    # its "top" side). Doing it the other way around (style, then merge)
    # silently discards the styling on every cell except the top-left --
    # confirmed by testing both orders directly, not assumed.
    ws.merge_cells(f"{first_col_letter}1:{last_col_letter}1")
    ws.merge_cells(f"{first_col_letter}2:{last_col_letter}2")

    for c in block_cols:
        cell1 = ws.cell(1, c)
        cell1.border = _SUMMARY_THIN_BORDER
        cell1.fill = _SUMMARY_SPOC_FILL
        cell2 = ws.cell(2, c)
        cell2.border = _SUMMARY_THIN_BORDER
        cell2.fill = _SUMMARY_DATE_FILL

    spoc_cell.font = _SUMMARY_SPOC_FONT
    spoc_cell.alignment = _SUMMARY_TOP_CENTER

    date_cell.font = _SUMMARY_SPOC_FONT
    date_cell.alignment = _SUMMARY_TOP_CENTER
    date_cell.number_format = "d-mmm-yy"

    for c, label in zip(block_cols, ("Surat", "Cuttack", "Durg", "Jharli", "Grand Total")):
        cell = ws.cell(3, c, label)
        cell.font = _SUMMARY_SPOC_FONT
        cell.alignment = _SUMMARY_TOP_CENTER
        cell.border = _SUMMARY_THIN_BORDER

    surat_l, cuttack_l, durg_l, jharli_l = (
        get_column_letter(c) for c in (surat_c, cuttack_c, durg_c, jharli_c)
    )
    for r in range(4, 15):
        # Literal 0, not blank -- every real block shows zero-counts as an
        # explicit "0", never an empty cell (confirmed against a real
        # corrected reference copy). Functionally identical to the Grand
        # Total formula either way (SUM treats blank as 0 too), but this
        # matches what Khagash's own blocks look like visually.
        for c in (surat_c, cuttack_c, durg_c, jharli_c):
            cell = ws.cell(r, c, 0)
            cell.font = _SUMMARY_SPOC_FONT
            cell.alignment = _SUMMARY_TOP_CENTER
            cell.border = _SUMMARY_THIN_BORDER
        total_cell = ws.cell(r, total_c, f"={jharli_l}{r}+{durg_l}{r}+{cuttack_l}{r}+{surat_l}{r}")
        total_cell.font = _SUMMARY_SPOC_FONT
        total_cell.alignment = _SUMMARY_TOP_CENTER
        total_cell.border = _SUMMARY_THIN_BORDER
    for c in block_cols:
        cell = ws.cell(15, c, 1)
        cell.font = _SUMMARY_SPOC_FONT
        cell.alignment = _SUMMARY_TOP_CENTER
        cell.number_format = "0%"
        cell.border = _SUMMARY_THIN_BORDER

    if new_start > 2:
        prev_start = new_start - 5
        for offset in range(5):
            prev_letter = get_column_letter(prev_start + offset)
            prev_dim = ws.column_dimensions.get(prev_letter)
            if prev_dim and prev_dim.width:
                ws.column_dimensions[get_column_letter(new_start + offset)].width = prev_dim.width
            # Collapse the block that was previously the newest -- matches
            # the real file's own convention (confirmed against a real
            # corrected reference copy): only the latest block stays
            # visible by default, older ones are folded away but still
            # there (un-hide manually in Excel if you need to look back).
            ws.column_dimensions[prev_letter].hidden = True


def _write_base_sheet(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Base", 0)
    columns = list(df.columns)

    ws.append(columns)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx)
        cell.font = BASE_FONT_NOT_BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = BASE_COL_WIDTHS.get(str(name).strip(), DEFAULT_COL_WIDTH)

    text_col_idxs = [i for i, name in enumerate(columns) if name in TEXT_FORMAT_COLUMNS]

    row_idx = 1
    for row in df.itertuples(index=False):
        row_idx += 1
        ws.append(row)
        for idx in text_col_idxs:
            ws.cell(row_idx, idx + 1).number_format = "@"
    return ws


# ---------------------------------------------------------------------------
# Entry point used by the generic upload -> process -> download route
# ---------------------------------------------------------------------------

def process(input_files: dict, dates: dict, output_dir: Path) -> Path:
    report_date_str = dates["report_date"]
    report_date = pd.Timestamp(report_date_str)
    log.info("Processing Control Tower Tracker for %s", report_date_str)

    prev_base = _load_prev_base(input_files["prev_tracker"])
    yct = _load_yesterday_completed_trips(input_files["daily_tracking_output"])
    package_sim_codes = _load_dump_code_set(input_files["package_sim_depot"])
    deviation_40km_codes = _load_dump_code_set(input_files["deviation_40km"])

    try:
        new_rows, stats = build_new_base_rows(
            prev_base, yct, package_sim_codes, deviation_40km_codes, report_date
        )
    except KeyError as exc:
        raise ReportProcessingError(
            f"Expected column {exc} not found. Check each file was uploaded to the correct slot "
            "(Previous Tracker / Daily Tracking Report / Package & SIM Depot / Deviation Upto 40 Km)."
        ) from exc

    log.info(
        "Raw candidates: %d | Dropped (already checked): %d | Removed by rule: %s | New rows: %d",
        stats["raw_candidates"], stats["dropped_already_checked"],
        stats["removed_by_rule"], stats["new_rows_added"],
    )

    full_base = pd.concat([prev_base, new_rows], ignore_index=True)

    # Base is small (tens to low hundreds of rows, unlike Report 6's tens of
    # thousands), so loading the previous file normally (to keep Summary
    # intact) is safe here -- no need for Report 6's read-only-values
    # workaround, that was specifically about a much bigger sheet.
    wb = load_workbook(input_files["prev_tracker"])
    if "Base" in wb.sheetnames:
        del wb["Base"]
    _write_base_sheet(wb, full_base)  # create_sheet(..., 0) already places it first
    _add_summary_date_block(wb, report_date)

    output_path = output_dir / f"Control_Tower_Tracker_{report_date_str}.xlsx"
    wb.save(output_path)
    return output_path


def process_dispatch(input_files: dict, dates: dict, output_dir: Path) -> Path:
    """Entry point wired into the registry below. Offloads to the office
    server over SSH when SSH_HOST is configured (see core/ssh_worker.py +
    office_server_worker.py at the repo root), so this report also runs on
    the office server's CPU/RAM instead of Render's. With no office server
    configured (the default), falls straight through to the same
    `process()` above unchanged -- no change to this report's actual logic
    or output, just where it executes.
    """
    from core.ssh_worker import is_configured, run_remote

    if is_configured():
        return run_remote("7", input_files, dates, output_dir)
    return process(input_files, dates, output_dir)


register(
    ReportMeta(
        id="7",
        name="Control Tower Tracker",
        input_slots=[
            InputSlot(
                key="prev_tracker",
                label="Previous Control Tower Tracker",
                accept=".xlsx",
                hint="Control_Tower_Tracker_<prev_date>.xlsx",
            ),
            InputSlot(
                key="daily_tracking_output",
                label="Daily Tracking Report Output (Report 3)",
                accept=".xlsx",
                hint="JKLC_Daily_Tracking_Report_<date>.xlsx",
            ),
            InputSlot(
                key="package_sim_depot",
                label="Package & SIM Depot List",
                accept=".xlsx",
                hint="Package & SIM Depot <month>.xlsx",
            ),
            InputSlot(
                key="deviation_40km",
                label="Deviation Upto 40 Km List",
                accept=".xlsx",
                hint="Deviation Upto 40 Km <month>.xlsx",
            ),
        ],
        output_pattern="Control_Tower_Tracker_<date>.xlsx (2 tabs: Base, Summary)",
        process_fn=process_dispatch,
        implemented=True,
        date_mode="single",
        notes=(
            "Stateful report -- needs the PREVIOUS Control Tower Tracker as an input each run. "
            "Candidates = Yesterday Completed Trips rows tagged 'Backward'; dedup by INVNO against "
            "all prior Base history; automatic removal rules (zone match, low Track Health, RMC site, "
            "Package/SIM Depot match, 40km allowance match) apply ONLY to still-'Backward' rows, never "
            "to human-reclassified 'Hit'/'Data Missed' rows. SOLD TO gets its leading D stripped on ANY "
            "D+letter prefix (confirmed broader than the original brief's 'DG/DD only' wording). HIT "
            "classification stays fully manual (confirmed no computable pattern). Validated exact "
            "against real data: 80/80 rows, identical INVNO sets, 0 unexpected diffs across all 111 "
            "columns. Summary tab is a manual data-entry table (per-date column blocks Khagash fills "
            "in by hand) -- carried forward, plus a new blank block scaffolded for this run's date "
            "(labels, Grand Total formula) ready for manual entry; the actual counts aren't computed."
        ),
    )
)
