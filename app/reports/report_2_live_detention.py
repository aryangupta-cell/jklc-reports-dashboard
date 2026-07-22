"""
JKLC Live Detention — web app adaptation (Report 2 of 9)
==========================================================

Adapted from the provided `jklc_live_detention.py` script. Business logic
is kept as given except for fixes confirmed against real data:

  1. `build_20km_candidates` — Step 5 date/end-time filter revised.
  2. `apply_detention_rules` — detention-hour rule now floor-only (no upper cap):
     keep rows with Detention Hours >= threshold, however long stuck.
  3. `SUMMARY_INCLUDE_REMARKS` — tightened to ["Detention"] only.

Validated against real 14 July data: 70/71 rows exact (1-row Durg gap
traced to a secondary inclusion rule not yet identified; not blocking).

Output is Summary-only (15 columns, formatted to match the real master
file: Calibri 11pt, centered, headers at row 3 starting col C). MTR and
20 KM tabs were tried (to match the real master file's 3-tab structure)
but removed again -- writing that much data (25k+ MTR rows x 121 cols,
x4 plant files) pushed peak memory to ~300MB+ on Render's free tier
(512MB limit) and caused an OOM-driven 502. Per explicit instruction,
staying Summary-only until the hosting plan is upgraded.

I/O differences from the original script:
  - Input: 3 uploaded files instead of local file paths.
  - Output: 4 plant-wise .xlsx files, zipped into one .zip.
  - Start/end dates come from the web form (manual range, matching Report 1's
    UI) instead of a module constant / auto-computed window.
"""

import logging
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reports.errors import ReportProcessingError, describe_column_mismatch
from reports.registry import InputSlot, ReportMeta, register

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONFIG — same values as the original script. Duplicated rather than
# imported from Report 1 (each report module stays fully self-contained,
# matching how the original standalone scripts were each independent).
# ---------------------------------------------------------------------------

MTR_DROP_COLUMNS = [
    "Probable Unloading Count",
    "Probable Unloading Detention",
    "1 Km Geofence Start Time",
    "1 Km Geofence End Time",
    "1 Km Geofence Detention",
    "20 Km Geofence Start Time",
    "20 Km Geofence End Time",
    "20 Km Geofence Detention",
    "40 Km Geofence Start Time",
    "40 Km Geofence End Time",
    "40 Km Geofence Detention",
    "Lap Sharable Link",
]

ORG_LOCATION_MAP = {
    "Durg Plant": "JKLC Durg",
    "Jharli Grinding": "JKLC Jharli",
    "JK Lakshmi Cement Limited- Surat Grinding Unit": "JKLC Surat",
    "JK Lakshmi Cement Limited - Surat Grinding Unit": "JKLC Surat",
    "MS JK LAKSHMI CEMENT LIMITED CUTTACK GRINDING UNIT": "JKLC Cuttack",
    "M/S JK Lakshmi Cement Limited Cuttack Grinding Unit": "JKLC Cuttack",
}

ORG_LOCATIONS_ALL = ["JKLC Cuttack", "JKLC Durg", "JKLC Jharli", "JKLC Surat"]

# CONFIRMED 14 July 2026 against real data (see notes below): only
# "Detention" belongs in the final Summary. Previous guess of
# ["Detention", "OUT OF GEOFENCE"] was wrong -- dropped.
SUMMARY_INCLUDE_REMARKS = ["Detention"]

# Exact raw column schema of the Detention Bot's output CSV (per the
# original process() error message, which already named these 4 columns
# explicitly): trip_id, remark, status, error.
BOT_OUTPUT_COLS = ["trip_id", "remark", "status", "error"]

SUMMARY_COLUMNS = [
    "Plant Name", "Invoice Number", "Quantity", "Distribution Channel",
    "Vehicle number", "Transporter", "Ship to Region", "Ship to District",
    "Unloading Point", "Invoice Date", "Detention Since", "Detention (Hours)",
    "Ship to Party", "Lead Distance", "Detention Days",
]


def _compute_date_window(report_date_str: str):
    """Confirmed: 1st of REPORT_DATE's month -> REPORT_DATE itself, NO -3
    day trim (different rule from Report 1). Confirmed both by the SOP text
    and the user's own reference table ("Current Month (1 to today)").

    Kept but NOT called from process() below — per explicit instruction, the
    UI takes a manual start/end range instead (consistent with Report 1),
    so this auto-compute isn't used as the default. Left in place in case a
    "suggest dates" convenience feature is wanted later."""
    report_date = pd.to_datetime(report_date_str)
    start = report_date.replace(day=1)
    end = report_date
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# File I/O
# ---------------------------------------------------------------------------

def _read_any(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            return pd.read_csv(path, dtype=str)
        if suffix in (".xlsx", ".xls"):
            return pd.read_excel(path, dtype=str)
    except Exception as exc:
        raise ReportProcessingError(f"Couldn't read '{path.name}' as a {suffix} file: {exc}") from exc
    raise ReportProcessingError(f"Unsupported file type '{suffix}' for '{path.name}'. Expected .csv or .xlsx.")


# ---------------------------------------------------------------------------
# Clean MTR (same as Report 1 / the original script — unchanged)
# ---------------------------------------------------------------------------

def clean_mtr(df: pd.DataFrame, start_date: str, end_date: str) -> pd.DataFrame:
    df = df.copy()
    df = df.drop(columns=[c for c in MTR_DROP_COLUMNS if c in df.columns])
    df["Org Location"] = df["Org Location"].replace(ORG_LOCATION_MAP)

    df["_inv_dt"] = pd.to_datetime(df["Invoice Date & Time"], errors="coerce")
    mask = (df["_inv_dt"] >= start_date) & (df["_inv_dt"] <= end_date + " 23:59:59")
    df = df[mask]

    return df


# ---------------------------------------------------------------------------
# Step 5: build the "20 KM" candidate sheet
#
# CHANGED from the original script — confirmed with the user 14 July 2026,
# based on direct verification against the real 13 July master output:
#
#   Original filter: Mode in {AT FIX, GPS-API}, Lead Distance > 20,
#   Proximity Start date == REPORT_DATE exactly, Proximity End Time blank.
#   The original script's docstring called this "CONFIRMED EXACT (173/173)."
#
#   Verification against the real file showed that claim was a coincidental
#   ROW-COUNT match, not a real row-identity match: comparing by Invoice No,
#   only 2 of 172 real rows actually matched. The other 170 real rows had
#   Proximity Start dates ranging from 6-12 July (detentions carrying over
#   across multiple days, not just "started today"), and 29/172 real rows
#   (17%) had a non-blank Proximity End Time yet were still included.
#
#   Revised filter: Proximity Start anywhere within [start_date, end_date]
#   (no exact-day requirement), End Time condition dropped entirely (Step 7's
#   Dispatch not-yet-delivered check is the real "still open" signal, not
#   Proximity End Time). This produces 95% overlap (164/172) with the real
#   20 KM sheet, up from ~1% with the original literal-SOP filter.
# ---------------------------------------------------------------------------

def build_20km_candidates(mtr_clean: pd.DataFrame, start_date: str, end_date: str) -> pd.DataFrame:
    df = mtr_clean.copy()
    df["Lead Distance"] = pd.to_numeric(df["Lead Distance"], errors="coerce")
    df["_proximity_start"] = pd.to_datetime(
        df["System Destination Proximity Start Time"], errors="coerce"
    )

    mask = (
        df["Mode"].isin(["AT FIX", "GPS-API"])
        & (df["Lead Distance"] > 20)
        & (df["_proximity_start"] >= start_date)
        & (df["_proximity_start"] <= end_date + " 23:59:59")
    )
    return df[mask].copy()


# ---------------------------------------------------------------------------
# Steps 6-9: Dispatch filter + detention-hour rule + Stamp Status rejected
# (unchanged from the original script)
# ---------------------------------------------------------------------------

def filter_dispatch_for_detention(dispatch_raw: pd.DataFrame) -> pd.DataFrame:
    df = dispatch_raw.copy()
    df["EPOD_Timestamp"] = pd.to_numeric(df["EPOD_Timestamp"], errors="coerce").fillna(0)
    df["MIGO_Timestamp"] = pd.to_numeric(df["MIGO_Timestamp"], errors="coerce").fillna(0)
    return df[(df["EPOD_Timestamp"] == 0) & (df["MIGO_Timestamp"] == 0)]


def apply_detention_rules(candidates: pd.DataFrame, dispatch_filtered: pd.DataFrame,
                           now: pd.Timestamp) -> pd.DataFrame:
    disp_inv = set(dispatch_filtered["Invoice Number"].astype(str).str.strip())
    df = candidates[candidates["Invoice no"].astype(str).str.strip().isin(disp_inv)].copy()

    # Detention (Hours) — NOT a raw MTR column, computed here (see original
    # script's docstring note on this being unconfirmed vs Khagash's exact
    # "NOW" reference point).
    df["Detention (Hours)"] = (now - df["_proximity_start"]).dt.total_seconds() / 3600

    # Detention-hour rule: floor only (no upper cap). Keep trips with
    # detention AT LEAST the threshold, however long they've been stuck.
    # Lead Distance ≤ 200 KM → floor is 24 hours.
    # Lead Distance > 200 KM → floor is 48 hours.
    mask = (
        ((df["Lead Distance"] <= 200) & (df["Detention (Hours)"] >= 24))
        | ((df["Lead Distance"] > 200) & (df["Detention (Hours)"] >= 48))
    )
    df = df[mask]

    df = df[df["Stamp Status"] != "Rejected"]
    return df


# ---------------------------------------------------------------------------
# Merge in the Detention Bot's remark (pass-through, no reclassification,
# unchanged from the original script)
# ---------------------------------------------------------------------------

def merge_bot_remarks(df: pd.DataFrame, bot_output: pd.DataFrame) -> pd.DataFrame:
    bot = bot_output[bot_output["status"] == "ok"][["trip_id", "remark"]].copy()
    bot["trip_id"] = bot["trip_id"].astype(str).str.strip()
    df = df.copy()
    df["Trip ID"] = df["Trip ID"].astype(str).str.strip()
    df = df.merge(bot, left_on="Trip ID", right_on="trip_id", how="inner")
    df = df.rename(columns={"remark": "Remark"})
    return df


# ---------------------------------------------------------------------------
# Build the Summary tab (unchanged from the original script)
# ---------------------------------------------------------------------------

def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # CONFIRMED 14 July 2026 -- see SUMMARY_INCLUDE_REMARKS definition above.
    df = df[df["Remark"].isin(SUMMARY_INCLUDE_REMARKS)]

    df["Detention Days"] = df["Detention (Hours)"] / 24

    out = pd.DataFrame({
        "Plant Name": df["Org Location"],
        "Invoice Number": df["Invoice no"],
        "Quantity": df.get("Quantity", ""),
        "Distribution Channel": df.get("Distribution Channel", ""),
        "Vehicle number": df["Vehicle No."],
        "Transporter": df["Transporter"],
        "Ship to Region": df.get("Ship to District", ""),
        "Ship to District": df.get("Ship to District", ""),
        "Unloading Point": df.get("Destination", ""),
        # Parsed to a real datetime (not the raw string) so Excel's
        # "m/d/yy h:mm" number format (matching the real file) actually
        # renders as a date instead of showing the literal text.
        "Invoice Date": pd.to_datetime(df["Invoice Date & Time"], errors="coerce"),
        "Detention Since": df["_proximity_start"],
        "Detention (Hours)": df["Detention (Hours)"],
        "Ship to Party": df.get("SOLD TO NM", ""),
        "Lead Distance": df["Lead Distance"],
        "Detention Days": df["Detention Days"],
    })
    return out


# ---------------------------------------------------------------------------
# Format utilities for Excel output — extracted cell-by-cell from the real
# master file (JKLC Live Detention Master 14th July 26.xlsx, Summary tab)
# so plant-wise files look identical to what Khagash produces by hand.
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", size=11, bold=False)
HEADER_FONT_BOLD = Font(name="Calibri", size=11, bold=True)  # "Detention Days" column only
DATA_FONT = Font(name="Calibri", size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# Header fill: real file uses theme color "accent2, lighter 40%" (computed
# from theme XML: ED7D31 accent2 + tint 0.6 -> F8CBAD) for every header cell
# except "Detention Days", which is bold with a plain yellow fill instead.
HEADER_FILL = PatternFill(fill_type="solid", fgColor="F8CBAD")
DETENTION_DAYS_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")

THIN = Side(style="thin")
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# Column widths, keyed by Summary column name (real file's C:Q widths).
# "Detention Since" has no explicit width in the real file (left at Excel's
# default ~8.43), so it's omitted here and falls through to the default below.
COLUMN_WIDTHS = {
    "Plant Name": 11.33, "Invoice Number": 14.0, "Quantity": 8.0,
    "Distribution Channel": 20.78, "Vehicle number": 13.66, "Transporter": 43.22,
    "Ship to Region": 14.66, "Ship to District": 22.78, "Unloading Point": 23.11,
    "Invoice Date": 15.44, "Detention (Hours)": 15.55, "Ship to Party": 41.33,
    "Lead Distance": 12.33, "Detention Days": 13.89,
}

# Real file formats Invoice Date / Detention Since as datetime, Detention
# Days as a whole number; everything else is left General.
DATETIME_COLUMNS = {"Invoice Date", "Detention Since"}
INTEGER_COLUMNS = {"Detention Days"}


def _write_summary_sheet(wb, df: pd.DataFrame):
    """Write the Summary sheet with headers at row 3, col C (matching the
    real master file): fills, borders, column widths, and number formats
    copied from the real file's Summary tab."""
    ws = wb.create_sheet("Summary")
    columns = list(df.columns)

    for col_idx, col_name in enumerate(columns):
        cell = ws.cell(3, 3 + col_idx, value=col_name)
        is_detention_days = col_name == "Detention Days"
        cell.font = HEADER_FONT_BOLD if is_detention_days else HEADER_FONT
        cell.fill = DETENTION_DAYS_FILL if is_detention_days else HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

        col_letter = ws.cell(3, 3 + col_idx).column_letter
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_name, 8.43)

    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        for col_idx, value in enumerate(row):
            col_name = columns[col_idx]
            cell = ws.cell(row_idx, 3 + col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            if col_name in DATETIME_COLUMNS:
                cell.number_format = "m/d/yy h:mm"
            elif col_name in INTEGER_COLUMNS:
                cell.number_format = "0"
    return ws


# ---------------------------------------------------------------------------
# Split into 4 plant-wise files, zip them for a single download
#
# NOTE: MTR and 20 KM tabs were tried (per an earlier request to match the
# real master file's 3-tab structure) but removed again -- writing that much
# data (25k+ MTR rows x 121 cols, x4 plant files) pushed peak memory to
# ~300MB+ on top of the interpreter/FastAPI baseline, which exceeds Render's
# free-tier 512MB limit and caused an OOM-driven 502. Per explicit
# instruction, staying Summary-only until the hosting plan can be upgraded.
# ---------------------------------------------------------------------------

def _write_plant_outputs_zip(summary: pd.DataFrame, date_label: str, output_dir: Path) -> Path:
    xlsx_paths = []

    for plant in ORG_LOCATIONS_ALL:
        plant_short = plant.replace("JKLC ", "")
        xlsx_path = output_dir / f"JKLC_Live_Detention_{plant_short}_{date_label}.xlsx"
        plant_summary = summary[summary["Plant Name"] == plant]

        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet
        _write_summary_sheet(wb, plant_summary)
        wb.save(xlsx_path)

        log.info("%s: %d rows -> %s", plant, len(plant_summary), xlsx_path.name)
        xlsx_paths.append(xlsx_path)

    zip_path = output_dir / f"JKLC_Live_Detention_{date_label}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in xlsx_paths:
            zf.write(p, arcname=p.name)
    return zip_path


# ---------------------------------------------------------------------------
# Entry point used by the generic upload -> process -> download route
# ---------------------------------------------------------------------------

def process(input_files: dict, dates: dict, output_dir: Path) -> Path:
    mtr_path = input_files["mtr_raw"]
    dispatch_path = input_files["dispatch"]
    bot_path = input_files["detention_bot"]

    # Manual start/end range, per explicit instruction (consistent with
    # Report 1) — NOT run through _compute_date_window(), which is kept but
    # unused (see that function's docstring for the confirmed "1st of month
    # to report date" rule it implements).
    start_date = dates["start_date"]
    end_date = dates["end_date"]
    log.info("Processing JKLC Live Detention for window %s to %s", start_date, end_date)

    mtr_raw = _read_any(mtr_path)
    dispatch_raw = _read_any(dispatch_path)
    bot_output = _read_any(bot_path)
    mismatch = describe_column_mismatch(bot_output.columns, BOT_OUTPUT_COLS, bot_path.name)
    if mismatch:
        raise ReportProcessingError(
            f"{mismatch} Check you uploaded the Detention Bot's output CSV."
        )

    try:
        mtr_clean = clean_mtr(mtr_raw, start_date, end_date)
        candidates = build_20km_candidates(mtr_clean, start_date, end_date)
        dispatch_filtered = filter_dispatch_for_detention(dispatch_raw)
    except KeyError as exc:
        raise ReportProcessingError(
            f"Expected column {exc} not found. Check you uploaded the correct "
            "file for each slot (MTR Raw / Daily Dispatch / Detention Bot output)."
        ) from exc

    now = pd.Timestamp(datetime.now())
    detained = apply_detention_rules(candidates, dispatch_filtered, now)

    try:
        merged = merge_bot_remarks(detained, bot_output)
    except KeyError as exc:
        raise ReportProcessingError(
            f"Expected column {exc} not found in the Detention Bot output. "
            "Check you uploaded the bot's output CSV (trip_id, remark, status, error)."
        ) from exc

    summary = build_summary(merged)
    log.info("Final Summary rows: %d (%s)", len(summary), dict(summary["Plant Name"].value_counts()))

    date_label = f"{start_date}_to_{end_date}"
    return _write_plant_outputs_zip(summary, date_label, output_dir)


def process_dispatch(input_files: dict, dates: dict, output_dir: Path) -> Path:
    """Entry point wired into the registry below. Offloads to the office
    server over SSH when SSH_HOST is configured (see core/ssh_worker.py +
    office_server_worker.py at the repo root), so this report also runs on
    the office server's CPU/RAM instead of Render's. With no office server
    configured (the default), falls straight through to the same
    `process()` above unchanged -- no change to this report's actual logic
    or output, just where it executes.
    """
    from core.ssh_worker import is_configured, run_remote

    if is_configured():
        return run_remote("2", input_files, dates, output_dir)
    return process(input_files, dates, output_dir)


register(
    ReportMeta(
        id="2",
        name="JKLC Live Detention",
        input_slots=[
            InputSlot(
                key="mtr_raw",
                label="MTR Raw",
                accept=".csv,.xlsx",
                hint="mtr - 2026-07-13T171021.301.csv",
            ),
            InputSlot(
                key="dispatch",
                label="Daily Dispatch",
                accept=".xlsx",
                hint="daily dispatch 13 Axestrack (75).xlsx",
            ),
            InputSlot(
                key="detention_bot",
                label="Detention Bot Output",
                accept=".csv",
                hint="bot output 13 detention_results.csv",
            ),
        ],
        output_pattern="JKLC_Live_Detention_<start>_to_<end>.zip (4 plant .xlsx files, Summary tab only)",
        process_fn=process_dispatch,
        implemented=True,
        date_mode="range",
        notes=(
            "Summary tab only (MTR/20 KM tabs tried but dropped -- OOM on Render free tier, "
            "see module docstring). Detention-hour rule is floor-only (no upper cap). "
            "Formatting matches real master file. Validated 70/71 rows on real 14 July data; "
            "1-row gap not root-caused."
        ),
    )
)
